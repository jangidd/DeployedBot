from collections import defaultdict
import math
from tkinter import Tk, filedialog
from urllib.parse import urlencode
from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse
from django.http import HttpResponse
from django.db import models, transaction
from django.contrib.auth.models import User
from django.contrib.auth.models import Group
from django.contrib.auth import authenticate
from django.contrib.auth import login as ContribLogin
from django.contrib.auth import logout as ContribLogout
from django.middleware.csrf import get_token
from django.urls import reverse
from users.models.AudiometryPdfReport import AudiometryReport
from users.models.OptometryPdfReport import OptometryReport
from users.models.VitalsPdfReport import VitalsReport
from users.models.instpersonalinfo import InstPersonalInfo as InstPersonalInfoModel
from users.models.institutionmodalities import InstitutionModalities as InstitutionModalitiesModel
from users.models.personalinfo import PersonalInfo as PersonalInfoModel
from users.models.qualificationdetails import QualificationDetails as QualificationDetailsModel
from users.models.workexp import WorkExp as WorkExpModel
from users.models.bankinginfo import BankingInfo as BankingInfoModel
from users.models.reportingarea import ReportingArea as ReportingAreaModel
from users.models.timeavailability import TimeAvailability as TimeAvailabilityModel
from users.models.patientdata import PatientInfo as PatientInfo
from users.models.patientdetails import PatientDetails as PatientDetails
from users.models.audiopatientdata import audioPatientDetails
from users.models.optometrydata import optopatientDetails
from users.models.vitalpatientdata import vitalPatientDetails
from users.models.VaccinationPatientData import vaccinationPatientDetails
from users.models.XRAYPatientData import xrayPatientDetails
from users.models.ECGPatientData import ecgPatientDetails
from users.models.DICOMData import DICOMData, DICOMFile, JPEGFile
from users.models.EcgPdfReport import EcgReport
from users.models.XrayPdfReport import XrayReport
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from users.forms import DICOMDataForm
from users.forms import PDFUploadForm
from users.forms import ECGUploadForm
from users.models.DailyCount import SetCount
from users.models.DailyCountECG import ECGSetCount
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.core.serializers import serialize
from django.contrib.auth import get_user_model
import json
import csv
from django.shortcuts import HttpResponse
from django.views import View
import os
import pydicom
from pydicom import dcmread
import matplotlib.pyplot as plt
from pydicom.data import get_testdata_files
from PIL import Image
from io import BytesIO
import pickle
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import io
from django.core.paginator import Paginator
from googleapiclient.http import MediaIoBaseDownload
import PyPDF2
from users.models.Date import Date
from datetime import datetime, timezone, timedelta
from users.models.Location import Location
from users.models.City import City
from users.models.Client import Client
from django.views.decorators.http import require_POST
from django.shortcuts import redirect
from functools import wraps
from .models.Xray_Client import XClient
from .models.Xray_City import XCity
from .models.Xray_Location import XLocation
from .models.Total_cases import Total_Cases
from django.db.models import F
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from django.contrib.sessions.models import Session
from django.contrib.sessions.backends.db import SessionStore
from django.contrib.auth.models import User
from django.utils import timezone
from django.contrib.auth import logout as contrib_logout
from django.contrib.auth.signals import user_logged_out, user_logged_in
from django.dispatch import receiver
from google.oauth2 import service_account
from google.oauth2.credentials import Credentials
from django.conf import settings
import fitz
import pandas as pd
from twilio.rest import Client as tw
import re
from django.utils.timezone import now

def login(request):
    if request.method == 'POST':
        email = request.POST['email']
        password = request.POST['password']

        user = authenticate(request, username=email, password=password)
        if user is not None:
            ContribLogin(request, user)
            group = user.groups.values_list('name', flat=True).first()

            if group == 'institution':
                return redirect('proinst')
            elif group == 'cardiologist':
                return redirect('ecgallocation')
            elif group == 'cardiologist2':
                return redirect('xrayallocation')
            elif group == 'audiometrist':
                return redirect('audiometry')
            elif group == 'ecgcoordinator':
                return redirect('ecgcoordinator')
            elif group == 'xraycoordinator':
                return redirect('xraycoordinator')
            elif group == 'technician':
                return redirect('upload_dicom')
            elif group == 'client':
                return redirect('client')
            elif group == 'campautomation':
                return redirect('optometrylist')
            else:
                return redirect('reportingbot')
        else:
            messages.add_message(request, messages.ERROR, "Invalid credentials")
            return render(request, 'users/login.html')
    return render(request, 'users/login.html')


def user_type_required(user_type):
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            if request.user.is_authenticated and request.user.groups.filter(name=user_type).exists():
                return view_func(request, *args, **kwargs)
            else:
                return redirect('login')

        return _wrapped_view

    return decorator

#@user_type_required('client')
def client_dashboard(request):
    try:
        current_user_personal_info = Client.objects.get(user=request.user)
    except Client.DoesNotExist:
        return HttpResponse("Client object does not exist for this user.", status=404)

    pdfs_list = []
    test_dates_set = set()
    report_dates_set = set()

    if current_user_personal_info.location:
        location = current_user_personal_info.location
        pdfs = XrayReport.objects.filter(location=location.name)  # Matching location name
        pdfs_list.extend(pdfs)
        test_dates_set.update(pdf.test_date for pdf in pdfs)
        report_dates_set.update(pdf.report_date for pdf in pdfs)

    formatted_test_dates = sorted(test_date.strftime('%Y-%m-%d') for test_date in test_dates_set)
    formatted_report_dates = sorted(report_date.strftime('%Y-%m-%d') for report_date in report_dates_set)

    context = {
        'pdfs': pdfs_list,
        'Test_Dates': formatted_test_dates,
        'Report_Dates': formatted_report_dates,
        'Location': current_user_personal_info.location
    }

    return render(request, 'users/client.html', context)

user_type_required('client')
def update_clinical_history(request):
    if request.method == 'GET':
        current_user_personal_info = XClient.objects.get(user=request.user)
        pdfs_list = []

        client = XClient.objects.filter(name__exact=current_user_personal_info).first()

        if client:
            cities = XCity.objects.filter(client=client)
            for city in cities:
                locations = XLocation.objects.filter(city=city)
                for location in locations:
                    print(location)
                    patient_data = DICOMData.objects.filter(location=location)
                    for data in patient_data:
                        pdfs_list.append({'patient_id': data.patient_id, 'patient_name': data.patient_name, 'clinical_history': data.notes})

        context = {
            'pdfs': pdfs_list,
        }

        return render(request, 'users/update_history.html', context)

    elif request.method == 'POST':
        data = json.loads(request.body)
        patient_id = data['patientId']
        new_clinical_history = data['newClinicalHistory']

        # Update clinical history in the database
        try:
            patient = DICOMData.objects.get(patient_id=patient_id)
            patient.notes = new_clinical_history
            patient.save()
            return JsonResponse({'success': True})
        except DICOMData.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Patient not found'})


def logout(request):
    contrib_logout(request)
    return redirect('login')


@user_type_required('ecgcoordinator')
#def allocation(request):
#    patients = PatientDetails.objects.all().order_by('-TestDate')  # Sort by date in ascending order
#    total_current_uploaded = PatientDetails.objects.all().count()
#    total_uploaded_ecg = Total_Cases.objects.values_list('total_uploaded_ecg', flat=True)
#    for value in total_uploaded_ecg:
#        total_uploaded_ecg = value

#    total_reported_ecg = Total_Cases.objects.values_list('total_reported_ecg', flat=True)
#    for value in total_reported_ecg:
#        total_reported_ecg = value

#    total_reported_patients = PatientDetails.objects.filter(cardiologist__isnull=False, isDone=True).count()
#    total_rejected_patients = PatientDetails.objects.filter(cardiologist__isnull=False, status=True).count()

#    total_unreported_and_unallocated_patients = PatientDetails.objects.filter(cardiologist=None, isDone=False).count()
#    total_unreported_and_allocated_patients = PatientDetails.objects.filter(cardiologist__isnull=False,
#                                                                            isDone=False).count()
#    total_unreported_patients = total_unreported_and_unallocated_patients + total_unreported_and_allocated_patients
#    total_cases = {'current_reported_cases': total_reported_patients, 'total_unreported': total_unreported_patients,
#                   'unallocated': total_unreported_and_unallocated_patients}
#    cardiologist_group = Group.objects.get(name='cardiologist')
#    cardiologists_objects = cardiologist_group.user_set.all()
#    unique_dates = set()
#    for patient in patients:
#        unique_dates.add(patient.date.date_field)
#    sorted_unique_dates = sorted(unique_dates, reverse=False)
#    formatted_dates = [date.strftime('%Y-%m-%d') for date in sorted_unique_dates]
#    unique_cities = [f"{x.name}" for x in City.objects.all()]
#    unique_locations = [f"{y.name}" for y in Location.objects.all()]
#    return render(request, 'users/allocation.html',
#                  {'total_cases': total_cases, 'total': total_current_uploaded, 'count': total_uploaded_ecg,
#                   'total_reported': total_reported_ecg, 'patients': patients, 'cardiologists': cardiologists_objects,
#                   'Date': formatted_dates, "Location": unique_locations, "Cities": unique_cities,
#                   "rejected": total_rejected_patients})

def allocation(request):
    success_message = ''
    success_details = []            
    rejected_message = ''
    rejected_details = []
    # Fetch and order patients
    patients = PatientDetails.objects.all().order_by('-TestDate')
    
    # Total counts for statistics
    total_current_uploaded = PatientDetails.objects.all().count()
    total_uploaded_ecg = Total_Cases.objects.values_list('total_uploaded_ecg', flat=True).first()
    total_reported_ecg = Total_Cases.objects.values_list('total_reported_ecg', flat=True).first()
    total_nonreported_ecg = Total_Cases.objects.values_list('total_nonreported_ecg', flat=True).first()

    total_reported_patients = PatientDetails.objects.filter(cardiologist__isnull=False, isDone=True).count()
    total_rejected_patients = PatientDetails.objects.filter(cardiologist__isnull=False, status=True).count()
    total_nonreported_patients = PatientDetails.objects.filter(NonReportable=True).count()
    total_unreported_and_unallocated_patients = PatientDetails.objects.filter(cardiologist=None, isDone=False).count()
    total_unreported_and_allocated_patients = PatientDetails.objects.filter(cardiologist__isnull=False, isDone=False).count()
    total_unreported_patients = total_unreported_and_unallocated_patients + total_unreported_and_allocated_patients

    total_cases = {
        'current_reported_cases': total_reported_patients,
        'total_unreported': total_unreported_patients,
        'unallocated': total_unreported_and_unallocated_patients,
        'nonreported': total_nonreported_patients
    }

    # Get cardiologists
    cardiologist_group = Group.objects.get(name='cardiologist')
    cardiologists_objects = cardiologist_group.user_set.all()

    # Set up pagination
    paginator = Paginator(patients, 400)  # 200 patients per page
    page_number = request.GET.get('page', 1)  # Get the page number from the request
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        page_obj = paginator.get_page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        page_obj = paginator.get_page(paginator.num_pages)

    # Get unique dates from the patients on the current page
    unique_dates = set(patient.date.date_field for patient in page_obj.object_list)
    sorted_unique_dates = sorted(unique_dates, reverse=False)
    formatted_dates = [date.strftime('%Y-%m-%d') for date in sorted_unique_dates]

    # Get unique cities and locations
    unique_cities = [f"{x.name}" for x in City.objects.all()]
    unique_locations = [f"{y.name}" for y in Location.objects.all()]

    # Everything related to the uploading of the ecg file to the database from the modal.
    form = ECGUploadForm()
    locations = Location.objects.all()

    # Handle messages from query parameters
    # success_message = request.GET.get('success_message', '')
    # success_details = request.GET.get('success_details', '')
    # rejected_message = request.GET.get('rejected_message', '')
    # rejected_details = request.GET.get('rejected_details', '')

    # if success_details:
    #     success_details = eval(success_details)  # Convert string representation of list back to list
    # if rejected_details:
    #     rejected_details = eval(rejected_details)  # Convert string representation of list back to list


    # success_message = ''
    # success_details = []
    # rejected_message = ''
    # rejected_details = []

    return render(request, 'users/allocation.html', {
        'total_cases': total_cases,
        'total': total_current_uploaded,
        'count': total_uploaded_ecg,
        'total_reported': total_reported_ecg,
        'total_nonreported': total_nonreported_ecg,
        'patients': page_obj,  # Use page_obj for pagination
        'cardiologists': cardiologists_objects,
        'Date': formatted_dates,
        'Location': unique_locations,
        'Cities': unique_cities,
        'rejected': total_rejected_patients,
        'page_obj': page_obj,  # Pass page_obj for pagination controls
        'location': locations,
        'form':form

    })  

@user_type_required('xraycoordinator')
#def allocation1(request):
#    patients = DICOMData.objects.all().order_by('-study_date')
#    total_current_uploaded = DICOMData.objects.all().count()

#    radiologist_group = Group.objects.get(name='cardiologist2')
#    radiologist_objects = radiologist_group.user_set.all()

#    total_uploaded_xray = Total_Cases.objects.values_list('total_uploaded_xray', flat=True)
#    for value in total_uploaded_xray:
#        total_uploaded_xray = value

#    total_reported_xray = Total_Cases.objects.values_list('total_reported_xray', flat=True)
#    for value in total_reported_xray:
#        total_reported_xray = value

#    total_nonreported_xray = Total_Cases.objects.values_list('total_nonreported_xray', flat=True)
#    for value in total_nonreported_xray:
#        total_nonreported_xray = value

#    total_reported_patients = DICOMData.objects.filter(isDone=True).count()
#    total_nonreported_patients = DICOMData.objects.filter(NonReportable=True).count()
#    total_unreported_and_unallocated_patients = DICOMData.objects.filter(radiologist=None, isDone=False, NonReportable=False).count()
#    total_unreported_and_allocated_patients = DICOMData.objects.filter(radiologist__isnull=False, isDone=False, NonReportable=False).values('patient_id').distinct().count()
#    total_unreported_patients = total_unreported_and_unallocated_patients + total_unreported_and_allocated_patients
#    total_cases = {'total_uploaded': total_uploaded_xray, 'alltime_reported': total_reported_xray,  'total_nonreported': total_nonreported_xray,
#                   'total_reported': total_reported_patients, 'total_unreported': total_unreported_patients,
#                   'unallocated': total_unreported_and_unallocated_patients, 'nonreported': total_nonreported_patients}
#    unique_dates = set()
#    for patient in patients:
#        unique_dates.add(patient.study_date)
#    sorted_unique_dates = sorted(unique_dates, reverse=False)
#    unique_locations = [f"{y.name}" for y in XLocation.objects.all()]
#    return render(request, 'users/allocation1.html',
#                  {"Location": unique_locations, 'total': total_cases, 'count': total_current_uploaded,
#                   'patients': patients, 'Date': sorted_unique_dates, 'radiologists': radiologist_objects})

def allocation1(request):
    # Fetch and order patients
    patients = DICOMData.objects.all().order_by('-study_date')
    
    # Total counts for statistics
    total_current_uploaded = DICOMData.objects.all().count()

    # Get radiologists from the group
    radiologist_group = Group.objects.get(name='cardiologist2')
    radiologist_objects = radiologist_group.user_set.all()

    # Retrieve total cases data
    total_uploaded_xray = Total_Cases.objects.values_list('total_uploaded_xray', flat=True).first()
    total_reported_xray = Total_Cases.objects.values_list('total_reported_xray', flat=True).first()
    total_nonreported_xray = Total_Cases.objects.values_list('total_nonreported_xray', flat=True).first()

    # Calculate various patient counts
    total_reported_patients = DICOMData.objects.filter(isDone=True).count()
    total_nonreported_patients = DICOMData.objects.filter(NonReportable=True).count()
    total_unreported_and_unallocated_patients = DICOMData.objects.filter(radiologist=None, isDone=False, NonReportable=False).count()
    total_unreported_and_allocated_patients = DICOMData.objects.filter(radiologist__isnull=False, isDone=False, NonReportable=False).values('patient_id').distinct().count()
    total_unreported_patients = total_unreported_and_unallocated_patients + total_unreported_and_allocated_patients

    total_cases = {
        'total_uploaded': total_uploaded_xray,
        'alltime_reported': total_reported_xray,
        'total_nonreported': total_nonreported_xray,
        'total_reported': total_reported_patients,
        'total_unreported': total_unreported_patients,
        'unallocated': total_unreported_and_unallocated_patients,
        'nonreported': total_nonreported_patients
    }

    # Set up pagination
    paginator = Paginator(patients, 400)  # 200 patients per page
    page_number = request.GET.get('page', 1)  # Get the page number from the request
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        page_obj = paginator.get_page(1)
    except EmptyPage:
        # If page is out of range, deliver last page of results
        page_obj = paginator.get_page(paginator.num_pages)

    # Get unique dates from the patients on the current page
    unique_dates = set(patient.study_date for patient in page_obj.object_list)
    sorted_unique_dates = sorted(unique_dates, reverse=False)

    # Get unique locations
    unique_locations = [f"{y.name}" for y in XLocation.objects.all()]

    return render(request, 'users/allocation1.html', {
        'Location': unique_locations,
        'total': total_cases,
        'count': total_current_uploaded,
        'patients': page_obj,
        'Date': sorted_unique_dates,
        'radiologists': radiologist_objects,
        'page_obj': page_obj  # Pass page_obj for pagination controls
    })


@user_type_required('ecgcoordinator')
def allocate(request):
    cardiologist_group = Group.objects.get(name='cardiologist')
    cardiologists_objects = cardiologist_group.user_set.all()

    total_unallocated_patients = PatientDetails.objects.filter(cardiologist=None, isDone=False)
    total_allocated_patients = PatientDetails.objects.filter(cardiologist__isnull=False, isDone=False)

    total_client = Client.objects.all()
    total_cities = City.objects.all()
    total_locations = Location.objects.all()
    total_dates = Date.objects.all()

    context = {
        'cardiologists': cardiologists_objects,
        'unallocated_patients': total_unallocated_patients,
        'allocated_patients': total_allocated_patients,
        'cities': total_cities,
        'clients': total_client,
        'locations': total_locations,
        'dates': total_dates,
    }

    if 'name' in request.POST:
        name = request.POST.get("name")
        email = request.POST.get("email")
        password = request.POST.get("password")

        client = Client(
            name=name,
            email=email,
            password=password,
        )
        client.save()
        return redirect("allocate")

    elif 'city_name' in request.POST:
        client_id = request.POST.get("client")
        city_name = request.POST.get("city_name")
        client = Client.objects.get(pk=client_id)
        city = City(client=client, name=city_name)
        city.save()

        return redirect("allocate")

    elif "location_name" in request.POST:
        city_id = request.POST.get('city')
        location_name = request.POST.get('location_name')
        city = City.objects.get(pk=city_id)
        location = Location(city=city, name=location_name)
        location.save()

        return redirect("allocate")


    elif 'delete_client' in request.POST:
        client_id = request.POST.get("delete_client")
        if client_id:
            client = Client.objects.filter(pk=client_id).first()
            if client:
                client.delete()

            return redirect("allocate")

    elif 'delete_city' in request.POST:
        city_id = request.POST.get("delete_city")
        if city_id:
            city = City.objects.filter(pk=city_id).first()
            if city:
                city.delete()

            return redirect("allocate")

    elif 'delete_location' in request.POST:
        location_id = request.POST.get("delete_location")
        if location_id:
            location = Location.objects.filter(pk=location_id).first()
            if location_id:
                location.delete()

            return redirect("allocate")

    action = request.POST.get('action')
    if action in ('allocate', 'unallocate','nonreport'):
        # Adding the non reportable button.
        if action == 'nonreport':
            selected_patient_ids = request.POST.getlist('cases')
            print(len(selected_patient_ids))
            if selected_patient_ids:
                # Fetching the patients who need to be updated and who are currently unreported.
                patients_to_update = PatientDetails.objects.filter(PatientId__in=selected_patient_ids, NonReportable=False)
                # Now updating the non reportable field to true.
                updated_nonreported_patients = patients_to_update.update(NonReportable=True)
                # Now updating the total ecg non reported cases only if the the number of true nonreportable patients are there.
                if updated_nonreported_patients > 0:
                    total_cases_instance = Total_Cases.objects.first()
                    total_cases_instance.total_nonreported_ecg += updated_nonreported_patients
                    total_cases_instance.save()
        # This was the original code before adding the nonreportable button. -Himanshu.
        else:
            selected_cardiologist_email = request.POST.get('cardiologist')
            if selected_cardiologist_email:
                cardiologist_group = Group.objects.get(name='cardiologist')
                cardiologist_user = get_object_or_404(cardiologist_group.user_set, email=selected_cardiologist_email)

                # Fetch the corresponding PersonalInfo instance for the selected cardiologist
                cardiologist = PersonalInfoModel.objects.get(user=cardiologist_user)

                if cardiologist:
                    selected_patient_ids = request.POST.getlist('cases')
                    if selected_patient_ids:
                        selected_patients = PatientDetails.objects.filter(PatientId__in=selected_patient_ids)
                        for patient in selected_patients:
                            if action == 'allocate' and patient.cardiologist != cardiologist:
                                patient.cardiologist = cardiologist
                                patient.save()
                            elif action == 'unallocate' and patient.cardiologist == cardiologist:
                                patient.cardiologist = None
                                patient.save()

    return render(request, 'users/allocate.html', context)


# Dictionary to store login and logout times for radiologists
radiologist_login_time = {}
radiologist_logout_time = {}


# Signal receiver for user login
@receiver(user_logged_in)
def handle_user_login(sender, request, user, **kwargs):
    login_time = timezone.now()
    radiologist_login_time[user.email] = login_time
    print('login time', radiologist_login_time)


# Signal receiver for user logout
@receiver(user_logged_out)
def handle_user_logout(sender, request, user, **kwargs):
    session_key = request.session.session_key
    session = Session.objects.get(session_key=session_key)
    logout_time = session.expire_date
    radiologist_logout_time[user.email] = logout_time
    print('logout time', radiologist_logout_time)


@user_type_required('xraycoordinator')
def allocate1(request):
    global radiologist_logout_time
    global radiologist_login_time

    radiologist_group = Group.objects.get(name='cardiologist2')
    radiologist_objects = radiologist_group.user_set.all()

    total_unallocated_patients = DICOMData.objects.filter(radiologist=None, isDone=False, NonReportable=False)
    total_allocated_patients = DICOMData.objects.filter(radiologist__isnull=False, isDone=False)
    total_reported_patients = DICOMData.objects.filter(isDone=True)



    total_client = XClient.objects.all()
    total_cities = XCity.objects.all()
    total_locations = XLocation.objects.all()
    total_dates = Date.objects.all()

    # Dictionary to store session status/first_name/time for each radiologist
    radiologist_session_status = {}

    for radiologist in radiologist_objects:
        is_active_session = False

        active_sessions = Session.objects.filter(expire_date__gte=timezone.now())
        for session in active_sessions:
            session_data = session.get_decoded()
            user_id = session_data.get('_auth_user_id')
            if user_id == str(radiologist.id):
                is_active_session = True
                break

        # Calculate inactive time for the current radiologist
        inactive_since_text = ""
        logout_time = radiologist_logout_time.get(radiologist.email)
        if logout_time:
            time_difference = timezone.now() - logout_time
            duration_seconds = time_difference.total_seconds()
            hours = int(duration_seconds // 3600)
            minutes = int((duration_seconds % 3600) // 60)
            seconds = int(duration_seconds % 60)

            if hours > 0:
                inactive_since_text = f"Inactive last {hours} hours"
            elif minutes > 0:
                inactive_since_text = f"Inactive last {minutes} minutes"
            else:
                inactive_since_text = f"Inactive last {seconds} seconds"

        active_since_text = ""
        login_time = radiologist_login_time.get(radiologist.email)
        if login_time:
            time_difference = timezone.now() - login_time
            duration_seconds = time_difference.total_seconds()
            hours = int(duration_seconds // 3600)
            minutes = int((duration_seconds % 3600) // 60)
            seconds = int(duration_seconds % 60)

            if hours > 0:
                active_since_text = f"Active last {hours} hours"
            elif minutes > 0:
                active_since_text = f"Active last {minutes} minutes"
            else:
                active_since_text = f"Active last {seconds} seconds"

        # Add email, first name, and inactive time to the radiologist_session_status dictionary
        radiologist_session_status[radiologist.email] = {
            'is_active': is_active_session,
            'first_name': radiologist.first_name,
            'inactive_time': inactive_since_text,
            'active_time': active_since_text
        }

    context = {
        'radiologist_session_status': radiologist_session_status,
        'unallocated_patients': total_unallocated_patients,
        'allocated_patients': total_allocated_patients,
        'reported_patients': total_reported_patients,
        'cities': total_cities,
        'clients': total_client,
        'locations': total_locations,
        'dates': total_dates,
    }

    if 'name' in request.POST:
        name = request.POST.get("name")
        email = request.POST.get("email")
        password = request.POST.get("password")

        client = XClient(
            name=name,
            email=email,
            password=password,
        )
        client.save()
        return redirect("allocate1")

    elif 'city_name' in request.POST:
        client_id = request.POST.get("client")
        city_name = request.POST.get("city_name")
        client = XClient.objects.get(pk=client_id)
        city = XCity(client=client, name=city_name)
        city.save()

        return redirect("allocate1")

    elif "location_name" in request.POST:
        city_id = request.POST.get('city')
        location_name = request.POST.get('location_name')
        city = XCity.objects.get(pk=city_id)
        location = XLocation(city=city, name=location_name)
        location.save()

        return redirect("allocate1")

    elif 'delete_client' in request.POST:
        client_id = request.POST.get("delete_client")
        if client_id:
            client = XClient.objects.filter(pk=client_id).first()
            if client:
                client.delete()

            return redirect("allocate1")

    elif 'delete_city' in request.POST:
        city_id = request.POST.get("delete_city")
        if city_id:
            city = XCity.objects.filter(pk=city_id).first()
            if city:
                city.delete()

            return redirect("allocate1")

    elif 'delete_location' in request.POST:
        location_id = request.POST.get("delete_location")
        if location_id:
            location = XLocation.objects.filter(pk=location_id).first()
            if location_id:
                location.delete()

            return redirect("allocate1")

    action = request.POST.get('action')
    if action in ('allocate', 'unallocate', 'nonreport'):
        if action == 'nonreport':
            selected_patient_ids = request.POST.getlist('cases')
            print(len(selected_patient_ids))
            if selected_patient_ids:
                # Fetching the patients filtered by comparing the id and nonreportable checkbox status.
                patients_to_update = DICOMData.objects.filter(patient_id__in=selected_patient_ids, NonReportable=False)
                # Only updating those who needs to be updated and are the nonreportable is not checked yet.
                updated_nonreported_patients = patients_to_update.update(NonReportable=True)
                # Now updating the total ecg non reported cases only if the the number of true nonreportable patients are there.
                if updated_nonreported_patients > 0:
                    total_cases_instance = Total_Cases.objects.first()
                    total_cases_instance.total_nonreported_xray += updated_nonreported_patients
                    total_cases_instance.save()

        else:
            selected_radiologist_email = request.POST.get('radiologist')
            if selected_radiologist_email:
                radiologist_group = Group.objects.get(name='cardiologist2')
                radiologist_user = get_object_or_404(radiologist_group.user_set, email=selected_radiologist_email)

                # Fetch the corresponding PersonalInfo instance for the selected cardiologist
                radiologist = PersonalInfoModel.objects.get(user=radiologist_user)

                if radiologist:
                    selected_patient_ids = request.POST.getlist('cases')
                    if selected_patient_ids:
                        selected_patients = DICOMData.objects.filter(patient_id__in=selected_patient_ids)

                        for patient in selected_patients:
                            if action == 'allocate':
                                patient.radiologist.add(radiologist)
                            elif action == 'unallocate':
                                patient.radiologist.remove(radiologist)

                    selected_patient_ids = request.POST.getlist('cases1')
                    if selected_patient_ids:
                        selected_patients = DICOMData.objects.filter(patient_id__in=selected_patient_ids)
                        for patient in selected_patients:
                            if action == 'allocate':
                                patient.radiologist.add(radiologist)
                            elif action == 'unallocate':
                                patient.radiologist.remove(radiologist)
                    selected_patient_ids = request.POST.getlist('cases2')
                    if selected_patient_ids:
                        selected_patients = DICOMData.objects.filter(patient_id__in=selected_patient_ids)
                        for patient in selected_patients:
                            if action == 'allocate':
                                patient.radiologist.add(radiologist)
                            elif action == 'unallocate':
                                patient.radiologist.remove(radiologist)
    return render(request, 'users/allocate1.html', context)


@user_type_required('cardiologist')
def ecgallocation(request):
    cardiologist_group = Group.objects.get(name='cardiologist')

    # Fetch the corresponding PersonalInfo instance for the current user
    current_user_personal_info = PersonalInfoModel.objects.get(user=request.user)
    total_reported = current_user_personal_info.total_reported
    today = now().date()
    yesterday = today - timedelta(days=1)

    allocated_to_current_user = PatientDetails.objects.filter(cardiologist=current_user_personal_info, isDone=False, status=False).order_by('TestDate')

    # Set up pagination
    paginator = Paginator(allocated_to_current_user, 200)  # 200 patients per page
    page_number = request.GET.get('page', 1)  # Get the page number from the request
    page_obj = paginator.get_page(page_number)

    # Get unique dates for patients in the current page
    unique_dates = set()
    for patient in page_obj.object_list:
        unique_dates.add(patient.date.date_field)
    sorted_unique_dates = sorted(unique_dates, reverse=False)
    formatted_dates = [date.strftime('%Y-%m-%d') for date in sorted_unique_dates]
    unique_location = Location.objects.all()

    return render(request, 'users/ecgallocation.html',
                  {
                      'reported': total_reported,
                      'patients': page_obj,
                      'Date': formatted_dates,
                      'Location': unique_location,
                      'page_obj': page_obj
                  })


@user_type_required('cardiologist2')
def xrayallocation(request):
    radiologist_group = Group.objects.get(name='cardiologist2')

    # Fetch the corresponding PersonalInfo instance for the current user
    current_user_personal_info = PersonalInfoModel.objects.get(user=request.user)
    total_reported = current_user_personal_info.total_reported
    today = now().date()
    yesterday = today - timedelta(days=1)
    allocated_to_current_user = DICOMData.objects.filter(radiologist=current_user_personal_info, isDone=False).order_by('study_date')

    # Set up pagination
    paginator = Paginator(allocated_to_current_user, 200)  # 200 patients per page
    page_number = request.GET.get('page', 1)  # Get the page number from the request
    page_obj = paginator.get_page(page_number)

    location = XLocation.objects.all()
    unique_dates = set()
    for patient in allocated_to_current_user:
        unique_dates.add(patient.study_date)
    sorted_unique_dates = sorted(unique_dates, reverse=False)
    return render(request, 'users/xrayallocation.html',
                  {'reported': total_reported, 'patients': page_obj, 'Date': sorted_unique_dates,
                   'locations': location, 'page_obj': page_obj})


@user_type_required('audiometrist')
def audiometry(request):
    patients = audioPatientDetails.objects.all()
    return render(request, 'users/audiometry.html', {'patients': patients})


def regrdo(request):
    return render(request, 'users/regrdo.html')


def reginst(request):
    return render(request, 'users/reginst.html')


@login_required
def prordo(request):
    return render(request, 'users/prordo.html')


@login_required
def proinst(request):
    return render(request, 'users/proinst.html')


# 1
def InstPersonalInfo(request):
    if request.method == 'POST':
        instfullname = request.POST['instfullname']
        instadd = request.POST['instadd']
        cnprname = request.POST['cnprname']
        cnprphone = request.POST['cnprphone']
        cnprdesignation = request.POST['cnprdesignation']
        altcnprname = request.POST['altcnprname']
        altcnprdesignation = request.POST['altcnprdesignation']
        altcnprphone = request.POST['altcnprphone']
        emailfrpacs = request.POST['emailfrpacs']
        emailfraccount = request.POST['emailfraccount']
        accountcnpr = request.POST['accountcnpr']
        acccnprphone = request.POST['acccnprphone']
        password1 = request.POST['password1']

        user = User.objects.create_user(username=emailfrpacs, email=emailfrpacs, password=password1,
                                        first_name=instfullname)

        insti_group = Group.objects.get(name="institution")
        insti_group.user_set.add(user)

        x = InstPersonalInfoModel.objects.create(user=user, instadd=instadd, cnprname=cnprname,
                                                 cnprphone=cnprphone,
                                                 cnprdesignation=cnprdesignation, altcnprname=altcnprname,
                                                 altcnprdesignation=altcnprdesignation, altcnprphone=altcnprphone,
                                                 emailfraccount=emailfraccount,
                                                 accountcnpr=accountcnpr,
                                                 acccnprphone=acccnprphone)

        x.save()
        print("Done.!!")
        return JsonResponse(status=201, data={"message": "success"})
    else:
        print("Not done..")
        return JsonResponse(status=400, data={"message": "invalid data"})


# 2
def InstitutionModalities(request):
    if request.method == 'POST':
        mriopt1 = ','.join(request.POST.getlist('mriopt1'))
        mriothers1 = request.POST['mriothers1']
        ctopt1 = ','.join(request.POST.getlist('ctopt1'))
        ctothers1 = request.POST['ctothers1']
        xray1 = True if request.POST.get('xray1') == 'on' else False
        others1 = True if request.POST.get('other1') == 'on' else False
        rdoprefrence = request.POST['rdoprefrence']
        exnocase = request.POST['exnocase']
        urgent = request.POST['urgent']
        nonurgent = request.POST['nonurgent']

        x = InstitutionModalitiesModel.objects.create(mriopt1=mriopt1, mriothers1=mriothers1, ctopt1=ctopt1,
                                                      ctothers1=ctothers1,
                                                      xray1=xray1, others1=others1,
                                                      rdoprefrence=rdoprefrence, exnocase=exnocase,
                                                      urgent=urgent, nonurgent=nonurgent)
        x.save()
        print("Done.!!")
        return JsonResponse(status=201, data={"message": "success", "redirect": True})
    else:
        print("Not done..")
        return JsonResponse(status=400, data={"message": "invalid data"})



User = get_user_model()


def PersonalInfo(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        password = request.POST.get('password')
        phone = request.POST.get('phone')
        altphone = request.POST.get('altphone')
        reference = request.POST.get('reference')
        resume = request.FILES.get('resume')
        uploadpicture = request.FILES.get('uploadpicture')
        signature = request.FILES.get('signature')
        companylogo = request.FILES.get('companylogo')
        serviceslist = request.POST.getlist('serviceslist')  # Get a list of selected services
        exportlist = request.POST.getlist('exportlist')

        if not all([name, email, password, phone, resume, uploadpicture, signature, companylogo, serviceslist,
                    exportlist]):
            return JsonResponse(status=400, data={"message": "Missing required fields"})

        user = User.objects.create_user(username=email, email=email, password=password, first_name=name)
        insti_group, _ = Group.objects.get_or_create(name="radiologist")
        insti_group.user_set.add(user)

        personal_info = PersonalInfoModel.objects.create(user=user, phone=phone, altphone=altphone,
                                                         reference=reference, resume=resume,
                                                         uploadpicture=uploadpicture, signature=signature,
                                                         companylogo=companylogo)
        personal_info.serviceslist.set(serviceslist)  # Set the ManyToManyField with the selected
        personal_info.exportlist.set(exportlist)

        return JsonResponse(status=201, data={"message": "success"})
    else:
        return JsonResponse(status=400, data={"message": "Invalid request method"})


# 4
def QualificationDetails(request):
    if request.method == 'POST':
        print(request.POST)
        tensname = request.POST['tensname']
        tengrade = request.POST['tengrade']
        tenpsyr = request.POST['tenpsyr']
        tencertificate = request.FILES['tencertificate']
        twelvesname = request.POST['twelvesname']
        twelvegrade = request.POST['twelvegrade']
        twelvepsyr = request.POST['twelvepsyr']
        twelvecertificate = request.FILES['twelvecertificate']
        mbbsinstitution = request.POST['mbbsinstitution']
        mbbsgrade = request.POST['mbbsgrade']
        mbbspsyr = request.POST['mbbspsyr']
        mbbsmarksheet = request.FILES['mbbsmarksheet']
        mbbsdegree = request.FILES['mbbsdegree']
        mdinstitution = request.POST['mdinstitution']
        mdgrade = request.POST['mdgrade']
        mdpsyr = request.POST['mdpsyr']
        mddegree = request.FILES['mddegree']

        x = QualificationDetailsModel.objects.create(tensname=tensname, tengrade=tengrade, tenpsyr=tenpsyr,
                                                     tencertificate=tencertificate,
                                                     twelvesname=twelvesname, twelvegrade=twelvegrade,
                                                     twelvepsyr=twelvepsyr, twelvecertificate=twelvecertificate,
                                                     mbbsinstitution=mbbsinstitution, mbbsgrade=mbbsgrade,
                                                     mbbspsyr=mbbspsyr,
                                                     mbbsmarksheet=mbbsmarksheet, mbbsdegree=mbbsdegree,
                                                     mdinstitution=mdinstitution, mdgrade=mdgrade, mdpsyr=mdpsyr,
                                                     mddegree=mddegree)
        x.save()
        print("Done.!!")
        return JsonResponse(status=201, data={"message": "success"})
    else:
        print("Not done..")
        return JsonResponse(status=400, data={"message": "invalid data"})


# 5
def WorkExp(request):
    if request.method == 'POST':
        print(request.POST)
        exinstitution = request.POST['exinstitution']
        exstdate = request.POST['exstdate']
        exenddate = request.POST['exenddate']
        designation = request.POST['designation']
        exinstitution1 = request.POST['exinstitution1']
        exstdate1 = request.POST['exstdate1']
        exenddate1 = request.POST['exenddate1']
        designation1 = request.POST['designation1']
        prexst = request.POST['prexst']
        prexend = request.POST['prexend']
        pii = request.POST['pii']
        msname = request.POST['msname']
        mcirgno = request.POST['mcirgno']
        regcecr = request.FILES['regcer']

        x = WorkExpModel.objects.create(exinstitution=exinstitution, exstdate=exstdate, exenddate=exenddate,
                                        designation=designation,
                                        exinstitution1=exinstitution1, exstdate1=exstdate1,
                                        exenddate1=exenddate1, designation1=designation1,
                                        prexst=prexst, prexend=prexend,
                                        pii=pii, msname=msname,
                                        mcirgno=mcirgno, regcecr=regcecr)
        x.save()
        print("Done.!!")
        return JsonResponse(status=201, data={"message": "success"})
    else:
        print("Not done..")
        return JsonResponse(status=400, data={"message": "invalid data"})


# 6
def BankingInfo(request):
    if request.method == 'POST':
        print(request.POST)
        bankname = request.POST['bankname']
        acnumber = request.POST['acnumber']
        ifsc = request.POST['ifsc']
        pancardno = request.POST['pancardno']
        pandcard = request.FILES['pancard']
        cheque = request.FILES['cheque']
        pictureproof = request.FILES['pictureproof']

        x = BankingInfoModel.objects.create(bankname=bankname, acnumber=acnumber, ifsc=ifsc,
                                            pancardno=pancardno,
                                            pandcard=pandcard, cheque=cheque,
                                            pictureproof=pictureproof)
        x.save()
        print("Done.!!")
        return JsonResponse(status=201, data={"message": "success"})
    else:
        print("Not done..")
        return JsonResponse(status=400, data={"message": "invalid data"})


# 7
def ReportingArea(request):
    if request.method == 'POST':
        print(request.POST)
        mriopt = ','.join(request.POST.getlist('mriopt'))
        mriothers = request.POST['mriothers']
        ctopt = ','.join(request.POST.getlist('ctopt'))
        ctothers = request.POST['ctothers']
        xray = True if request.POST.get('xray') == 'on' else False
        others = True if request.POST.get('other') == 'on' else False

        x = ReportingAreaModel.objects.create(mriopt=mriopt, mriothers=mriothers, ctopt=ctopt,
                                              ctothers=ctothers,
                                              xray=xray, others=others)
        x.save()
        print("Done.!!")
        return JsonResponse(status=201, data={"message": "success"})
    else:
        print("Not done..")
        return JsonResponse(status=400, data={"message": "invalid data"})


# 8
def TimeAvailability(request):
    if request.method == 'POST':
        print(request.POST)
        monday = True if request.POST.get('monday') == 'on' else False
        tuesday = True if request.POST.get('tuesday') == 'on' else False
        wednesday = True if request.POST.get('wednesday') == 'on' else False
        thursday = True if request.POST.get('thursday') == 'on' else False
        friday = True if request.POST.get('friday') == 'on' else False
        saturday = True if request.POST.get('saturday') == 'on' else False
        sunday = True if request.POST.get('sunday') == 'on' else False
        monst = request.POST.get('monst')
        monend = request.POST.get('monend')
        tuest = request.POST.get('tuest')
        tueend = request.POST.get('tueend')
        wedst = request.POST.get('wedst')
        wedend = request.POST.get('wedend')
        thust = request.POST.get('thust')
        thuend = request.POST.get('thuend')
        frist = request.POST.get('frist')
        friend = request.POST.get('friend')
        satst = request.POST.get('satst')
        satend = request.POST.get('satend')
        sunst = request.POST.get('sunst')
        sunend = request.POST.get('sunend')

        x = TimeAvailabilityModel.objects.create(monday=monday, tuesday=tuesday, wednesday=wednesday,
                                                 thursday=thursday,
                                                 friday=friday, saturday=saturday, sunday=sunday, monst=monst,
                                                 monend=monend, tuest=tuest, tueend=tueend, wedst=wedst, wedend=wedend,
                                                 thust=thust, thuend=thuend, frist=frist, friend=friend, satst=satst,
                                                 satend=satend, sunst=sunst, sunend=sunend)
        x.save()
        print("Done.!!")
        return JsonResponse(status=201, data={"message": "success", "redirect": True})
    else:
        print("Not done..")
        return JsonResponse(status=400, data={"message": "invalid data"})


@csrf_exempt
def userExists(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email__exact=email)
        except User.DoesNotExist:
            user = None

        if (user is not None):
            return JsonResponse(status=200, data="This email has already been taken", safe=False)
        else:
            return JsonResponse(status=200, data=user is None, safe=False)


# Create your views here.

@csrf_exempt
def numberExists(request):
    if request.method == 'POST':
        cnprphone = request.POST.get('cnprphone')
        try:
            x = InstPersonalInfoModel.objects.get(cnprphone__exact=cnprphone)
        except InstPersonalInfoModel.DoesNotExist:
            x = None

        if (x is not None):
            return JsonResponse(status=200, data="This phone number already exist", safe=False)
        else:
            return JsonResponse(status=200, data=x is None, safe=False)


@csrf_exempt
def phoneExists(request):
    if request.method == 'POST':
        phone = request.POST.get('phone')
        try:
            x = PersonalInfoModel.objects.get(phone__exact=phone)
        except PersonalInfoModel.DoesNotExist:
            x = None

        if (x is not None):
            return JsonResponse(status=200, data="This phone number already exist", safe=False)
        else:
            return JsonResponse(status=200, data=x is None, safe=False)


# *************************************************** CSV Upload for General Purpose *******************************************************************
@csrf_exempt
def patientData(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = PatientInfo.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId__icontains=query) | Q(PatientName__icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)


# Added by Aman at 05:46

# def uploadcsv(request):
#     if request.method == 'POST' and request.FILES['csv_file']:
#         csv_file = request.FILES['csv_file']
#
#         # Adjust the field names according to your CSV file structure
#         field_names = ['PatientId', 'PatientName', 'age', 'gender', 'TestDate', 'ReportDate', 'height', 'weight',
#                        'blood', 'pulse', 'FarVisionRight', 'FarVisionLeft', 'NearVisionRight', 'NearVisionLeft',
#                        'ColorBlindness']
#
#         try:
#             # Decode the CSV file data and split it into lines
#             decoded_file = csv_file.read().decode('utf-8').splitlines()
#
#             # Parse the CSV data using the DictReader
#             reader = csv.DictReader(decoded_file, fieldnames=field_names)
#
#             # Skip the header row if it exists
#             if reader.fieldnames == field_names:
#                 next(reader)
#
#             # Iterate over each row and insert into the PatientInfo table
#             for row in reader:
#                 # Convert date strings to datetime objects if not blank
#                 test_date = datetime.strptime(row['TestDate'], '%d-%m-%Y').date() if row['TestDate'] else None
#                 report_date = datetime.strptime(row['ReportDate'], '%d-%m-%Y').date() if row['ReportDate'] else None
#
#                 # Convert datetime objects back to strings in the desired format if not None
#                 test_date_formatted = test_date.strftime('%Y-%m-%d') if test_date else None
#                 report_date_formatted = report_date.strftime('%Y-%m-%d') if report_date else None
#
#                 PatientInfo.objects.create(
#                     PatientId=row['PatientId'],
#                     PatientName=row['PatientName'],
#                     age=row['age'],
#                     gender=row['gender'],
#                     TestDate=test_date_formatted,
#                     ReportDate=report_date_formatted,
#                     height=row['height'],
#                     weight=row['weight'],
#                     blood=row['blood'],
#                     pulse=row['pulse'],
#                     FarVisionRight=row['FarVisionRight'],
#                     FarVisionLeft=row['FarVisionLeft'],
#                     NearVisionRight=row['NearVisionRight'],
#                     NearVisionLeft=row['NearVisionLeft'],
#                     ColorBlindness=row['ColorBlindness'],
#                 )
#
#             return HttpResponse('CSV file uploaded successfully.')
#         except Exception as e:
#             return HttpResponse(f'Error: {str(e)}')
#     else:
#         # return HttpResponse('Please upload a CSV file.')
#         return render(request, 'users/uploadcsv.html')


# audiometry****************************************************************** CSV Upload ***************************************************************************


def audiopatientDetails(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = audioPatientDetails.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId__icontains=query) | Q(PatientName__icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)


def uploadcsvforaudio(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']

        # Adjust the field names according to your CSV file structure
        field_names = ['Name', 'Patient ID', 'Age', 'Gender', 'TestDate', 'ReportDate',
                       'Left Air conduction DB 1 (250 Hz)',
                       'Left Air conduction DB 2 (500 Hz)', 'Left Air conduction DB 3 (1000 Hz)',
                       'Left Air conduction DB 4 (2000 Hz)', 'Left Air conduction DB 5 (4000 Hz)',
                       'Left Air conduction DB 6 (8000 Hz)', 'Left Bone Conduction 1 (250 Hz)',
                       'Left Bone Conduction 2 (500 Hz)', 'Left Bone Conduction 3 (1000 Hz)',
                       'Left Bone Conduction 4 (2000 Hz)', 'Left Bone Conduction 5 (4000 Hz)',
                       'Right Air Conduction 1 (250 Hz)', 'Right Air Conduction 2 (500 Hz)',
                       'Right Air Conduction 3 (1000 Hz)', 'Right Air Conduction 4 (2000 Hz)',
                       'Right Air Conduction 5 (4000 Hz)', 'Right Air Conduction 6 (8000 Hz)',
                       'Right Bone Conduction 1 (250 Hz)', 'Right Bone Conduction 2 (500 Hz)',
                       'Right Bone Conduction 3 (1000 Hz)', 'Right Bone Conduction 4 (2000 Hz)',
                       'Right Bone Conduction 5 (4000 Hz)', 'Left Ear Finding', 'Right Ear Finding']

        try:
            # Decode the CSV file data and split it into lines
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file, fieldnames=field_names)
            if reader.fieldnames == field_names:
                next(reader)

            # Initialize a list to store missing data logs
            missing_data_logs = []
            total_rows = 0
            saved_rows = 0

            # Store the CSV data in a variable
            csv_data = list(reader)

            # Check for missing data in each row
            for idx, row in enumerate(csv_data, start=1):
                total_rows += 1
                missing_fields = [field for field in field_names if not row.get(field)]
                if missing_fields:
                    # Append each missing data message separately for each row
                    error_message = f"Missing data for ID: {row.get('PatientId')} and Name: {row.get('PatientName')} in row {idx}: {', '.join(missing_fields)}"
                    missing_data_logs.append(error_message)
                    messages.error(request, error_message)
                else:
                    saved_rows += 1
                    # Extract values for Left Air conduction DB columns
                    left_ear_db_values = [
                        row['Left Air conduction DB 1 (250 Hz)'],
                        row['Left Air conduction DB 2 (500 Hz)'],
                        row['Left Air conduction DB 3 (1000 Hz)'],
                        row['Left Air conduction DB 4 (2000 Hz)'],
                        row['Left Air conduction DB 5 (4000 Hz)'],
                        row['Left Air conduction DB 6 (8000 Hz)'],
                    ]

                    # Concatenate values with commas and store in leftEarDB field
                    left_ear_db_combined = ', '.join(left_ear_db_values)

                    # Extract values for Left Bone conduction DB columns
                    left_ear_bone_db_values = [
                        row['Left Bone Conduction 1 (250 Hz)'],
                        row['Left Bone Conduction 2 (500 Hz)'],
                        row['Left Bone Conduction 3 (1000 Hz)'],
                        row['Left Bone Conduction 4 (2000 Hz)'],
                        row['Left Bone Conduction 5 (4000 Hz)'],
                    ]

                    # Concatenate values with commas and store in leftEarDB field
                    left_ear_bone_db_combined = ', '.join(left_ear_bone_db_values)

                    # Extract values for Right Air conduction DB columns
                    right_ear_db_values = [
                        row['Right Air Conduction 1 (250 Hz)'],
                        row['Right Air Conduction 2 (500 Hz)'],
                        row['Right Air Conduction 3 (1000 Hz)'],
                        row['Right Air Conduction 4 (2000 Hz)'],
                        row['Right Air Conduction 5 (4000 Hz)'],
                        row['Right Air Conduction 6 (8000 Hz)'],
                    ]

                    # Concatenate values with commas and store in leftEarDB field
                    right_ear_db_combined = ', '.join(right_ear_db_values)

                    # Extract values for Right Bone conduction DB columns
                    right_ear_bone_db_values = [
                        row['Right Bone Conduction 1 (250 Hz)'],
                        row['Right Bone Conduction 2 (500 Hz)'],
                        row['Right Bone Conduction 3 (1000 Hz)'],
                        row['Right Bone Conduction 4 (2000 Hz)'],
                        row['Right Bone Conduction 5 (4000 Hz)'],
                    ]

                    # Concatenate values with commas and store in leftEarDB field
                    right_ear_bone_db_combined = ', '.join(right_ear_bone_db_values)

                    # Convert date strings to datetime objects
                    test_date = datetime.strptime(row['TestDate'], '%d-%m-%Y').date()
                    report_date = datetime.strptime(row['ReportDate'], '%d-%m-%Y').date()

                    # Convert datetime objects back to strings in the desired format
                    test_date_formatted = test_date.strftime('%Y-%m-%d')
                    report_date_formatted = report_date.strftime('%Y-%m-%d')

                    audioPatientDetails.objects.create(
                        PatientId=row['Patient ID'],
                        PatientName=row['Name'],
                        age=row['Age'],
                        gender=row['Gender'],
                        TestDate=test_date_formatted,
                        ReportDate=report_date_formatted,
                        leftEarDB=left_ear_db_combined,
                        leftEarBoneDB=left_ear_bone_db_combined,
                        rightEarDB=right_ear_db_combined,
                        rightEarBoneDB=right_ear_bone_db_combined,
                        rightEarLevel=row['Left Ear Finding'],
                        leftEarLevel=row['Right Ear Finding'],
                    )

            if missing_data_logs:
                # Include total rows and saved rows in the error message
                error_message = f'\nTotal rows: {total_rows}, Saved rows: {saved_rows}'
                messages.error(request, error_message)
                return redirect('audiometry')
            else:
                messages.success(request, 'CSV data uploaded successfully.')
                return redirect('audiometry')

        except Exception as e:
            return HttpResponse(f'Error: {str(e)}')
    else:
        # return HttpResponse('Please upload a CSV file.')
        return render(request, 'users/uploadcsv.html')


# optometry****************************************************************** CSV Upload ***************************************************************************
def optopatientDetails(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = optopatientDetails.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId__icontains=query) | Q(PatientName__icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)


def uploadcsvforopto(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']

        # Adjust the field names according to your CSV file structure
        field_names = ['Timestamp', 'Name', 'Patient ID', 'Age', 'Gender', 'Far vision right', 'Far vision left',
                       'Near vision right', 'Near vision left', 'Colour vision']

        try:
            # Decode the CSV file data and split it into lines
            decoded_file = csv_file.read().decode('utf-8').splitlines()

            # Parse the CSV data using the DictReader
            reader = csv.DictReader(decoded_file, fieldnames=field_names)

            # Skip the header row if it exists
            if reader.fieldnames == field_names:
                next(reader)

            # Iterate over each row and insert into the PatientInfo table
            for row in reader:

                # Extract date and time from Timestamp
                timestamp_str = row['Timestamp']
                print(timestamp_str)

                try:
                    # Try parsing with seconds included
                    timestamp_datetime = datetime.strptime(timestamp_str, '%d-%m-%Y %H:%M:%S')

                except ValueError:
                    # If parsing with seconds fails, try without seconds
                    timestamp_datetime = datetime.strptime(timestamp_str, '%m/%d/%Y %H:%M')

                    # Extract only the date part and format it as day/month/year
                timestamp_date = timestamp_datetime.date()
                timestamp_date_str = timestamp_date.strftime('%Y-%m-%d')

                optopatientDetails.objects.create(
                    PatientId=row['Patient ID'],
                    PatientName=row['Name'],
                    age=row['Age'],
                    gender=row['Gender'],
                    TestDate=timestamp_date_str,
                    ReportDate=timestamp_date_str,
                    FarVisionRight=row['Far vision right'],
                    FarVisionLeft=row['Far vision left'],
                    NearVisionRight=row['Near vision right'],
                    NearVisionLeft=row['Near vision left'],
                    ColorBlindness=row['Colour vision'],
                )

            return HttpResponse('CSV file uploaded successfully.')
        except Exception as e:
            return HttpResponse(f'Error: {str(e)}')
    else:
        # return HttpResponse('Please upload a CSV file.')
        return render(request, 'users/uploadcsv.html')


# vital****************************************************************** CSV Upload ***************************************************************************
def vitalpatientDetails(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = vitalPatientDetails.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId__icontains=query) | Q(PatientName__icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)



def fetch_patient_data(request):
    patient_id = request.GET.get('patientId')
    patient_name = request.GET.get('patientName')
    age = request.GET.get('age')
    gender = request.GET.get('gender')
    HeartRate = request.Get.get('HeartRate')
    test_date = request.GET.get('testDate')
    report_date = request.GET.get('reportDate')
    report_image = request.GET.get('reportImage')  # Report image URL

    # You can modify this logic based on how you fetch patient data
    patient = request(
        PatientDetails,
        PatientId=patient_id,
        PatientName=patient_name,
        age=age,
        gender=gender,
        HeartRate=HeartRate,
        TestDate=test_date,
        ReportDate=report_date,
        reportimage=report_image
    )

    # Create a dictionary to hold the patient data
    patient_data = {
        'PatientId': patient.PatientId,
        'PatientName': patient.PatientName,
        'age': patient.age,
        'gender': patient.gender,
        'HeartRate': patient.HeartRate,
        'TestDate': patient.TestDate.strftime('%Y-%m-%d'),  # Format the date as needed
        'ReportDate': patient.ReportDate.strftime('%Y-%m-%d'),  # Format the date as needed
        'reportimage': patient.reportimage.url,  # Get the URL of the report image
    }

    return JsonResponse(patient_data)
    # Check if the patient's report status is updated


class Google(View):
    def get(self, request):
        location = request.GET.get('location')
        date = request.GET.get('date')
        google_drive_data = GoogleDrive(location, date)
        response = HttpResponse(google_drive_data)
        return response


def GoogleDrive(location, date):
    # The file that contains the OAuth 2.0 credentials.
    CLIENT_SECRET_FILE = 'users/GoogleDriveAPI.json'

    # The name of the API and version of the API.
    API_NAME = 'drive'
    API_VERSION = 'v3'

    # The scopes that are required to access the API.
    SCOPES = ['https://www.googleapis.com/auth/drive']

    def create_service():
        # Create the credentials.
        creds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)

        # If the credentials don't exist or are invalid, then create new ones.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                # Create the flow object.
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)

                # Run the flow to obtain the credentials.
                creds = flow.run_local_server(port=0)

                # Save the credentials for future use.
                with open('token.pickle', 'wb') as token:
                    pickle.dump(creds, token)
        # Create the service object.
        service = build(API_NAME, API_VERSION, credentials=creds)
        return service

    # folder_id = '1RjxYJcv4vbv1WFfcUtCWm-qh0N3KRd0n'
    folder_id = '1DweRTm3gIqnbbYQxi-0gCtO-0IDLyTY1'
    service = create_service()

    existing_patient_ids = set(PatientDetails.objects.values_list('PatientId', flat=True))
    fetch_patient_data_from_folder(service, folder_id, existing_patient_ids, location, date)

def fetch_patient_data_from_folder(service, folder_id, existing_patient_ids, selected_location, date):
    stack = [(folder_id, None)]
    patient_data_by_date = {}

    while stack:
        current_folder_id, current_location_id = stack.pop()
        query = f"'{current_folder_id}' in parents and mimeType='application/vnd.google-apps.folder'"
        subfolders = service.files().list(q=query).execute()

        for subfolder in subfolders.get('files', []):
            subfolder_id = subfolder['id']
            subfolder_name = subfolder['name']
            print(f"Folder Name: {subfolder_name}, ID: {subfolder_id}")

            stack.append((subfolder_id, subfolder_name))
            technician_name = subfolder_name
            location_id = None

            if technician_name:
                location = Location.objects.filter(technician_name=technician_name).first()
                if location:
                    location_id = location.id

                query = f"'{subfolder_id}' in parents"
                subfolder_files = service.files().list(q=query).execute().get('files', [])

                for data in subfolder_files:
                    if data['mimeType'] == 'application/pdf':
                        file_id = data['id']
                        request = service.files().get_media(fileId=file_id)
                        pdf_files = io.BytesIO()
                        downloader = MediaIoBaseDownload(pdf_files, request)
                        done = False
                        while not done:
                            status, done = downloader.next_chunk()
                        pdf_reader = PyPDF2.PdfReader(pdf_files)

                        for page_number, page in enumerate(pdf_reader.pages):
                            first_page_text = page.extract_text()

                            id = str(first_page_text).split("Id")[1].split("\n")[0]
                            if ":" in id:
                                patient_id = id.split(":")[1].strip()
                            else:
                                patient_id = id.strip()

                            if patient_id == '':
                                if first_page_text.count('Comments') > 1:
                                    patient_id = str(first_page_text).split("Comments\nComments")[1].split("HR")[0].split('\n')[1].split('\n')[0]
                                else:
                                    patient_id = str(first_page_text).split("Comments")[1].split("HR")[0].strip()

                            if patient_id not in existing_patient_ids and patient_id != '':
                                if not PatientDetails.objects.filter(PatientId=patient_id).exists():
                                    name = str(first_page_text).split("Name")[1].split("\n")[0].split(":")[1].strip()
                                    if name == '':
                                        patient_name = 'None'
                                    else:
                                        patient_name = name

                                    patient_age = str(first_page_text).split("Age")[1].split("\n")[0].split(":")[1].strip()
                                    if patient_age == '':
                                        patient_age = 0
                                    if 'Gender' in first_page_text:
                                        patient_gender = str(first_page_text).split("Gender")[1].split("\n")[0].split(":")[1].strip()
                                    else:
                                        patient_gender = 'Missing'
                                    heart_rate = str(first_page_text).split("HR:")[1].split(" ")[1].split("/")[0].strip()
                                    pr_interval = str(first_page_text).split("PR:")[1].split("QRS:")[0].split("ms")[0].strip()
                                    report_time = str(first_page_text).split("Acquired on:")[1][12:17].strip()
                                    raw_date = str(first_page_text).split("Acquired on:")[1][0:11].strip()
                                    formatted_date = datetime.strptime(raw_date, '%Y-%m-%d').date()

                                    if formatted_date not in patient_data_by_date:
                                        patient_data_by_date[formatted_date] = []
                                    patient_data_by_date[formatted_date].append((patient_id, patient_name, patient_age, patient_gender, int(heart_rate),int(pr_interval), report_time))

                                    reportimage = os.path.join('ecg_graphs', f"{patient_id}_{patient_name}.pdf")

                                    if location_id is not None:
                                        existing_patient_ids.add(patient_id)

                                        date, created = Date.objects.get_or_create(date_field=formatted_date,
                                                                                   location_id=location_id)

                                        patient = PatientDetails(
                                            PatientId=patient_id,
                                            PatientName=patient_name,
                                            age=patient_age,
                                            gender=patient_gender,
                                            HeartRate=heart_rate,
                                            PRInterval=pr_interval,
                                            TestDate=formatted_date,
                                            ReportDate=formatted_date,
                                            date_id=date.id,
                                            reportimage=reportimage,
                                            location=location
                                        )
                                        patient.save()
                                        print(f"Patient saved: {patient}")
                                        total_cases, created = Total_Cases.objects.get_or_create(id=1, defaults={
                                            'total_uploaded_ecg': 0})
                                        total_cases.total_uploaded_ecg += 1
                                        total_cases.save()

                                        # Convert PDF page to image using PyMuPDF
                                        try:
                                            images = []
                                            pdf_stream = io.BytesIO(pdf_files.getvalue())
                                            doc = fitz.open(stream=pdf_stream, filetype='pdf')
                                            page = doc.load_page(page_number)
                                            image_bytes = page.get_pixmap().tobytes()
                                            image_buffer = io.BytesIO(image_bytes)
                                            image_file = ContentFile(image_buffer.getvalue(),
                                                                     name=f"{patient_id}_{patient_name}.jpg")
                                            # Save image data to the image field
                                            patient.image = image_file
                                            patient.save()
                                        except Exception as e:
                                            print(f"Error converting PDF page to image: {e}")

                                        try:
                                            pdf_filename = f"{patient_id}_{patient_name}.pdf"
                                            pdf_filepath = os.path.join('ecg_graphs', pdf_filename)
                                            with open(pdf_filepath, 'wb') as pdf_file:
                                                pdf_file.write(pdf_files.getvalue())
                                        except Exception as e:
                                            print(f"Error saving PDF file: {e}")




def report_patient(request, patient_id):
    # Your reporting logic here

    # Set a session variable to indicate that the button has been reported
    request.session[f"reportButtonState_{patient_id}"] = "reported"

    return HttpResponse("Reported successfully")  # You can customize the response as needed


# data***********************************
@csrf_exempt
def patientDetails(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = PatientDetails.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId__icontains=query) | Q(PatientName__icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)


# def upload_dicom(request):
#     form = DICOMDataForm()
#     locations = XLocation.objects.all()
#
#
#     success_message = ''
#     success_details = []
#     rejected_message = ''
#     rejected_details = []
#
#     if request.method == 'POST':
#         print(request.POST)
#         form = DICOMDataForm(request.POST, request.FILES)
#
#         location_name = request.POST.get('location')
#         locations = XLocation.objects.filter(name=location_name)
#
#         if locations.exists():  # Check if any locations were found
#             location = locations.first()  # Get the first location from the queryset
#             city = location.city  # Assuming XLocation has a ForeignKey to XCity
#             client = city.client  # Assuming XCity has a ForeignKey to XClient
#
#             if form.is_valid():
#                 print("Form is valid!")
#                 dicom_instances = []
#                 rejected_files = []
#
#                 for dicom_file in request.FILES.getlist('dicom_file'):
#
#                     try:
#                         dicom_data = dcmread(dicom_file)
#                     except Exception as e:
#                         print(f"Error reading DICOM file: {str(e)}")
#                         rejected_files.append({'id': None, 'name': dicom_file.name})
#                         continue
#
#                     study_date_formatted = datetime.strptime(dicom_data.StudyDate, "%Y%m%d").strftime("%Y-%m-%d")
#
#                     existing_instance = DICOMData.objects.filter(
#                         patient_id=str(dicom_data.PatientID),
#                     ).first()
#
#                     if existing_instance:
#                         print(f"Skipping file {dicom_file.name} - Duplicate data found.")
#                         rejected_files.append({'id': existing_instance.id, 'name': dicom_file.name})
#                     else:
#                         print(f"Saving file {dicom_file.name}")
#                         dicom_instance = DICOMData(dicom_file=dicom_file)
#                         dicom_instance.client = client
#                         dicom_instance.city = city
#                         dicom_instance.location = location
#                         dicom_instances.append(dicom_instance)
#
#                         # Save the instance to get the DICOM file path
#                         dicom_instance.save()
#
#                         # Extract metadata
#                         dicom_instance.patient_id = str(dicom_data.PatientID)
#                         dicom_instance.patient_name = str(dicom_data.PatientName)
#                         dicom_instance.age = str(dicom_data.PatientAge)
#
#                         dicom_instance.gender = 'Male' if dicom_data.PatientSex.upper() == 'M' else 'Female'
#                         dicom_instance.notes = request.POST.get("note")
#                         if dicom_instance.notes == '':
#                             dicom_instance.notes = "No Clinical history."
#
#                         # Format the study_date as "date/month/year"
#                         if dicom_data.StudyDate:
#                             datetime_obj = datetime.strptime(dicom_data.StudyDate, "%Y%m%d")
#                             dicom_instance.study_date = datetime_obj.strftime("%Y-%m-%d")
#                         else:
#                             dicom_instance.study_date = None
#
#                         dicom_instance.study_description = str(dicom_data.StudyDescription)
#
#                         # Convert DICOM image to JPEG-compatible format
#                         pixel_data = dicom_data.pixel_array
#                         if dicom_data.BitsAllocated == 16:
#                             pixel_data = pixel_data.astype('uint16')  # Convert to 16-bit unsigned integer
#                             pixel_data = pixel_data >> (dicom_data.BitsStored - 8)  # Right-shift to 8-bit
#
#                         # Convert DICOM image to JPEG and save
#                         with BytesIO() as output:
#                             Image.fromarray(pixel_data).convert('L').save(output, format='JPEG')  # 'L' for grayscale
#
#                             # Save the JPEG file
#                             dicom_instance.jpeg_file.save(f"{dicom_data.SOPInstanceUID}.jpg",
#                                                           ContentFile(output.getvalue()))
#
#                         # Update and save the instance with metadata
#                         dicom_instance.save()
#                         print(f"Total DICOM instances to save: {len(dicom_instances)}")
#
#                 # After the loop ends
#                 total_cases, created = Total_Cases.objects.get_or_create(id=1, defaults={'total_uploaded_xray': 0})
#
#                 if dicom_instances:
#                     total_cases.total_uploaded_xray += len(dicom_instances)
#                     success_message = f"{len(dicom_instances)} Images uploaded successfully."
#                     success_details = [{'id': None, 'name': dicom_instance.dicom_file.name} for dicom_instance in
#                                        dicom_instances]
#
#                 if rejected_files:
#                     rejected_message = f"{len(rejected_files)} files were rejected. Please check and try again."
#                     rejected_details = [{'id': item['id'], 'name': item['name']} for item in rejected_files]
#
#                 total_cases.save()
#
#         else:
#             print("No location found with the name:", location_name)
#
#     # Print both messages at the end
#     print("Success Message:", success_message)
#     print("Rejected Message:", rejected_message)
#
#     # Handle GET requests separately
#     return render(request, 'users/upload_dicom.html', {
#         'form': form,
#         'location': locations,
#         'success_message': success_message,
#         'success_details': success_details,
#         'rejected_message': rejected_message,
#         'rejected_details': rejected_details,
#     })

@user_type_required('technician')
def upload_dicom(request):
    form = DICOMDataForm()
    locations = XLocation.objects.all()

    success_message = ''
    success_details = []
    rejected_message = ''
    rejected_details = []

    if request.method == 'POST':
        form = DICOMDataForm(request.POST, request.FILES)
        upload_type = request.POST.get('upload_type')

        if upload_type == 'single_file_per_person':
            return handle_single_file_per_person_upload(request, form, locations)
        elif upload_type == 'multiple_file_single_person':
            return handle_multiple_file_single_person_upload(request, form, locations)
        else:
            return HttpResponse("Invalid upload type")

    return render(request, 'users/upload_dicom.html', {
        'form': form,
        'location': locations,
        'success_message': success_message,
        'success_details': success_details,
        'rejected_message': rejected_message,
        'rejected_details': rejected_details,
    })


def handle_single_file_per_person_upload(request, form, locations):
    success_message = ''
    rejected_message = ''
    success_details = []
    rejected_details = []
    if request.method == 'POST':
        form = DICOMDataForm(request.POST, request.FILES)

        location_name = request.POST.get('location')
        locations = XLocation.objects.filter(name=location_name)

        if locations.exists():
            location = locations.first()
            city = location.city
            client = city.client

            if form.is_valid():
                print("Form is valid!")
                dicom_instances = []
                rejected_files = []

                for dicom_file in request.FILES.getlist('dicom_file'):
                    try:
                        dicom_data = dcmread(dicom_file)
                        print(dicom_data)
                    except Exception as e:
                        print(f"Error reading DICOM file: {str(e)}")
                        rejected_files.append({'id': None, 'name': dicom_file.name})
                        continue

                    study_date_formatted = datetime.strptime(dicom_data.StudyDate, "%Y%m%d").strftime("%Y-%m-%d")
                    # Extract body part examined from DICOM file

                    accession_number = dicom_data.get('AccessionNumber', None)
                    if accession_number == '':
                        accession_number = None

                    with transaction.atomic():
                        existing_instance = DICOMData.objects.filter(
                            patient_id=str(dicom_data.PatientID),
                            # Add other relevant fields for comparison
                        ).first()

                        if existing_instance:
                            print(f"Skipping file {dicom_file.name} - Duplicate data found.")
                            rejected_files.append({'id': existing_instance.id, 'name': dicom_file.name})
                        else:
                            print(f"Saving file {dicom_file.name}")

                            dicom_instance = DICOMData.objects.create(
                                patient_name=str(dicom_data.PatientName),
                                patient_id=str(dicom_data.PatientID),
                                age=str(dicom_data.PatientAge),
                                gender='Male' if dicom_data.PatientSex.upper() == 'M' else 'Female',
                                study_date=study_date_formatted,
                                study_description=str(dicom_data.StudyDescription),
                                notes=request.POST.get("note"),
                                body_part_examined=str(dicom_data.BodyPartExamined),
                                location=location,
                                accession_number=accession_number
                            )
                            if dicom_instance.notes == '':
                                dicom_instance.notes = 'No Clinical History.'

                            # Save the DICOM file
                            dicom_file_obj = DICOMFile.objects.create(
                                dicom_data=dicom_instance,
                                dicom_file=dicom_file
                            )

                            # Convert DICOM image to JPEG-compatible format
                            pixel_data = dicom_data.pixel_array
                            if dicom_data.BitsAllocated == 16:
                                pixel_data = pixel_data.astype('uint16')
                                pixel_data = pixel_data >> (dicom_data.BitsStored - 8)

                            # Convert DICOM image to JPEG and save
                            with BytesIO() as output:
                                Image.fromarray(pixel_data).convert('L').save(output, format='JPEG')

                                # Save the JPEG file with the correct DICOM instance
                                jpeg_file_name = f"{dicom_file.name.split('.')[0]}.jpg"
                                jpeg_file = ContentFile(output.getvalue(), name=jpeg_file_name)
                                jpeg_instance = JPEGFile.objects.create(dicom_data=dicom_instance, jpeg_file=jpeg_file)
                                dicom_instance.save()

                            dicom_instances.append(dicom_instance)

                            print(f"Total DICOM instances to save: {len(dicom_instances)}")

                # After the loop ends
                total_cases, created = Total_Cases.objects.get_or_create(id=1, defaults={'total_uploaded_xray': 0})

                #Auto Allocate
                radiologist_group = Group.objects.get(name='cardiologist2')
                radiologist_user = get_object_or_404(radiologist_group.user_set, email='drgauravbpl@gmail.com')
                radiologist = PersonalInfoModel.objects.get(user=radiologist_user)

                selected_patient_id = DICOMData.objects.filter(location=17)
                for patient in selected_patient_id:
                    patient.radiologist.add(radiologist)

                if dicom_instances:
                    total_cases.total_uploaded_xray += len(dicom_instances)
                    success_message = f"{len(dicom_instances)} Images uploaded successfully."
                    success_details = [
                        {'id': dicom_instance.id, 'name': dicom_instance.dicom_files.first().dicom_file.name} for
                        dicom_instance in dicom_instances]


                if rejected_files:
                    rejected_message = f"{len(rejected_files)} files were rejected. Please check and try again."
                    rejected_details = [{'id': item['id'], 'name': item['name']} for item in rejected_files]

                total_cases.save()

        else:
            print("No location found with the name:", location_name)

    # Print both messages at the end
    print("Success Message:", success_message)
    print("Rejected Message:", rejected_message)

    # Handle GET requests separately
    return render(request, 'users/upload_dicom.html', {
        'form': form,
        'location': locations,
        'success_message': success_message,
        'success_details': success_details,
        'rejected_message': rejected_message,
        'rejected_details': rejected_details,
    })


def handle_multiple_file_single_person_upload(request, form, locations):
    success_message = ''
    rejected_message = ''
    success_details = []
    rejected_details = []
    if request.method == 'POST':
        form = DICOMDataForm(request.POST, request.FILES)
        upload_type = request.POST.get('upload_type')

        location_name = request.POST.get('location')
        locations = XLocation.objects.filter(name=location_name)
        body_part_examined = ''

        if locations.exists():
            location = locations.first()
            city = location.city
            client = city.client

            if form.is_valid():
                print("Form is valid!")
                dicom_instances = defaultdict(list)
                rejected_files = []

                for dicom_file in request.FILES.getlist('dicom_file'):
                    try:
                        dicom_data = dcmread(dicom_file)
                        print(dicom_data)
                    except Exception as e:
                        print(f"Error reading DICOM file: {str(e)}")
                        rejected_files.append({'id': None, 'name': dicom_file.name})
                        continue

                    study_date_formatted = datetime.strptime(dicom_data.StudyDate, "%Y%m%d").strftime("%Y-%m-%d")

                    accession_number = dicom_data.get('AccessionNumber', None)
                    if accession_number == '':
                        accession_number = None

                    # Get the unique identifier for the patient
                    patient_id = str(dicom_data.PatientID)

                    # Get the body part examined
                    if dicom_data.BodyPartExamined:
                        body_part_examined = str(dicom_data.BodyPartExamined.split()[0])


                    # Check if DICOMData instance already exists for the patient and study date
                    existing_instance = DICOMData.objects.filter(patient_id=patient_id).first()

                    if existing_instance and existing_instance.body_part_examined != body_part_examined:
                        # If an instance exists and body part examined is different, modify the patient ID
                        suffix = 1
                        while DICOMData.objects.filter(patient_id=f"{patient_id}-{suffix}").exists():
                            suffix += 1
                        patient_id = f"{patient_id}-{suffix}"

                    # Create or get DICOMData instance
                    dicom_instance, created = DICOMData.objects.get_or_create(
                        patient_id=patient_id,
                        body_part_examined=body_part_examined,
                        # Add other relevant fields for comparison
                    )

                    if created:
                        # Initialize common fields if this is a newly created DICOMData instance
                        dicom_instance.client = client
                        dicom_instance.city = city
                        dicom_instance.location = location
                        dicom_instance.patient_name = str(dicom_data.PatientName)
                        # if
                        dicom_instance.age = str(dicom_data.PatientAge)
                        dicom_instance.gender = 'Male' if dicom_data.PatientSex.upper() == 'M' else 'Female'
                        dicom_instance.notes = request.POST.get("note")
                        if dicom_instance.notes == '':
                            dicom_instance.notes = "No Clinical History."

                        dicom_instance.study_description = str(dicom_data.StudyDescription)  # Save Body Part Examined
                        dicom_instance.study_date = study_date_formatted
                        dicom_instance.accession_number = accession_number
                        dicom_instance.save()

                    # Create a DICOMFile instance for the DICOMData instance
                    dicom_file_obj = DICOMFile.objects.create(dicom_data=dicom_instance, dicom_file=dicom_file)

                    # Convert DICOM image to JPEG-compatible format
                    pixel_data = dicom_data.pixel_array
                    if dicom_data.BitsAllocated == 16:
                        pixel_data = pixel_data.astype('uint16')  # Convert to 16-bit unsigned integer
                        pixel_data = pixel_data >> (dicom_data.BitsStored - 8)  # Right-shift to 8-bit

                    # Convert DICOM image to JPEG and save
                    with BytesIO() as output:
                        Image.fromarray(pixel_data).convert('L').save(output, format='JPEG')  # 'L' for grayscale

                        # Save the JPEG file with the correct DICOM instance
                        jpeg_file_name = f"{dicom_file.name.split('.')[0]}.jpg"  # Assuming DICOM file name is unique
                        jpeg_file = ContentFile(output.getvalue(), name=jpeg_file_name)
                        jpeg_instance = JPEGFile.objects.create(dicom_data=dicom_instance, jpeg_file=jpeg_file)

                    # Keep track of successfully processed instances
                    if patient_id not in dicom_instances:
                        dicom_instances[patient_id] = dicom_instance

                # Retrieving total_cases
                total_cases, created = Total_Cases.objects.get_or_create(id=1, defaults={'total_uploaded_xray': 0})

                # Update total cases count
                total_cases.total_uploaded_xray += len(dicom_instances)
                total_cases.save()

                # Collect success details
                success_message = f"{len(dicom_instances)} Images uploaded successfully."
                success_details = [{'id': None, 'name': file_obj.dicom_file.name} for dicom_instance in
                                   dicom_instances.values() for file_obj in dicom_instance.dicom_files.all()]
                if rejected_files:
                    rejected_message = f"{len(rejected_files)} files were rejected. Please check and try again."
                    rejected_details = [{'id': item['id'], 'name': item['name']} for item in rejected_files]

        else:
            print("No location found with the name:", location_name)

    # Print both messages at the end
    print("Success Message:", success_message)
    print("Rejected Message:", rejected_message)

    # Handle GET requests separately
    return render(request, 'users/upload_dicom.html', {
        'form': form,
        'location': locations,
        'success_message': success_message,
        'success_details': success_details,
        'rejected_message': rejected_message,
        'rejected_details': rejected_details,
    })


@require_POST
def update_patient_done_status(request, patient_id):
    try:
        current_user_personal_info = PersonalInfoModel.objects.get(user=request.user)
        total_unreported_and_allocated_patients = PatientDetails.objects.filter(cardiologist=current_user_personal_info,
                                                                                isDone=False).count()

        if total_unreported_and_allocated_patients > 0:
            PersonalInfoModel.objects.filter(id=current_user_personal_info.id).update(
                total_reported=F('total_reported') + 1)

        total_uploaded_ecg = Total_Cases.objects.values_list('total_uploaded_ecg', flat=True)
        for value in total_uploaded_ecg:
            total_uploaded_ecg = value

        total_reported_ecg = Total_Cases.objects.values_list('total_reported_ecg', flat=True)
        for value in total_reported_ecg:
            total_reported_ecg = value

        if total_uploaded_ecg > total_reported_ecg:
            Total_Cases.objects.update(total_reported_ecg=F('total_reported_ecg') + 1)
        patient = PatientDetails.objects.get(PatientId=patient_id)
        patient.isDone = True
        patient.save()

        return JsonResponse({'success': True})
    except PatientDetails.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Patient not found'})


@require_POST
def update_patient_done_status_xray(request, patient_id):
    try:
        current_user_personal_info = PersonalInfoModel.objects.get(user=request.user)
        user_unreported_and_allocated_patients = DICOMData.objects.filter(radiologist=current_user_personal_info,
                                                                          isDone=False).count()
        print(user_unreported_and_allocated_patients)
        if user_unreported_and_allocated_patients > 0:
            PersonalInfoModel.objects.filter(id=current_user_personal_info.id).update(
                total_reported=F('total_reported') + 1)

        total_uploaded_xray = Total_Cases.objects.values_list('total_uploaded_xray', flat=True)
        for value in total_uploaded_xray:
            total_uploaded_xray = value

        total_reported_xray = Total_Cases.objects.values_list('total_reported_xray', flat=True)
        for value in total_reported_xray:
            total_reported_xray = value

        if total_uploaded_xray > total_reported_xray:
            Total_Cases.objects.update(total_reported_xray=models.F('total_reported_xray') + 1)
        patient = get_object_or_404(DICOMData, patient_id=patient_id)
        patient.isDone = True
        patient.save()

        return JsonResponse({'success': True})
    except DICOMData.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Patient not found'})


############################## ECG PDF upload to portal ###################################################
@user_type_required('ecgcoordinator')
def handle_ecg_pdf_upload(pdf_file):
    your_pdf_model_instance = EcgReport(pdf_file=pdf_file)
    your_pdf_model_instance.save()


def upload_ecg_pdf(request):
    if request.method == 'POST':
        try:
            print("Inside upload_pdf view")
            pdf_file = request.FILES.get('pdf')
            patient_id = request.POST.get('patientId')
            patient_name = request.POST.get('patientName')
            location = request.POST.get('location')
            test_date_str = request.POST.get('testDate')
            report_date_str = request.POST.get('reportDate')

            if not pdf_file:
                return JsonResponse({'error': 'No PDF file provided.'}, status=400)

            # Specify the upload path and create a folder if it doesn't exist
            upload_path = os.path.join('uploads', 'ecg_pdfs')
            os.makedirs(upload_path, exist_ok=True)

            # Save the PDF file to the specified path
            pdf_file_path = os.path.join(upload_path, pdf_file.name)
            print("PDF file path:", pdf_file_path)

            with open(pdf_file_path, 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)

            # Convert report_date_str to a datetime object
            test_date = datetime.strptime(test_date_str, "%Y-%m-%d").date()
            report_date = datetime.strptime(report_date_str, "%Y-%m-%d").date()

            # Save the PDF file path and additional data to the database
            pdf_model_instance = EcgReport(
                pdf_file=pdf_file_path,
                name=patient_name,
                patient_id=patient_id,
                location=location,
                test_date=test_date,
                report_date=report_date,

            )
            pdf_model_instance.save()

            return JsonResponse({'message': 'PDF successfully uploaded and processed.'})
        except Exception as e:
            print("Error processing PDF:", e)
            return JsonResponse({'error': 'Internal server error.'}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=400)


def get_csrf_token(request):
    csrf_token = {
        'csrf_token': get_token(request),

    }
    print("CSRF Token from headers:", csrf_token)
    return JsonResponse(csrf_token)


@user_type_required('ecgcoordinator')
############################ To retrive data
def ecg_pdf_report(request):
    pdfs = EcgReport.objects.all().order_by('-report_date')
    # Collect unique dates and locations from the PDFs
    test_dates = set(pdf.test_date for pdf in pdfs)
    formatted_dates = [date.strftime('%Y-%m-%d') for date in test_dates]
    report_dates = set(pdf.report_date for pdf in pdfs)
    unique_locations = set(pdf.location for pdf in pdfs)

    context = {
        'pdfs': pdfs,
        'Test_Date': sorted(formatted_dates),
        'Report_Date': sorted(report_dates),  # Ensure dates are sorted for dropdown
        'Location': sorted(unique_locations),  # Ensure locations are sorted for dropdown
    }

    return render(request, 'users/ecg_pdf_report.html', context)


############################################################# End of ECG PDF upload ###################################################################33

################################################################### XRAY PDF upload to portal #########################################################################
@user_type_required('xraycoordinator')
def handle_xray_pdf_upload(pdf_file):
    your_pdf_model_instance = XrayReport(pdf_file=pdf_file)
    your_pdf_model_instance.save()


def upload_xray_pdf(request):
    if request.method == 'POST':
        try:
            print("Inside upload_xray_pdf view")
            pdf_file = request.FILES.get('pdf')
            patient_id = request.POST.get('patientId')
            patient_name = request.POST.get('patientName')
            location = request.POST.get('location')
            accession_number = request.POST.get('accession')
            test_date_str = request.POST.get('testDate')
            report_date_str = request.POST.get('reportDate')

            if not pdf_file:
                return JsonResponse({'error': 'No PDF file provided.'}, status=400)

            # Specify the upload path and create a folder if it doesn't exist
            upload_path = os.path.join('uploads', 'xray_pdfs')
            os.makedirs(upload_path, exist_ok=True)
  
            

            # Save the PDF file to the specified path
            pdf_file_path = os.path.join(upload_path, pdf_file.name)
            print("PDF file path:", pdf_file_path)

            with open(pdf_file_path, 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)

            # Convert report_date_str to a datetime object
            test_date = datetime.strptime(test_date_str, "%Y-%m-%d").date()
            report_date = datetime.strptime(report_date_str, "%Y-%m-%d").date()

            # Save the PDF file path and additional data to the database
            pdf_model_instance = XrayReport(
                pdf_file=pdf_file_path,
                name=patient_name,
                patient_id=patient_id,
                location=location,
                test_date=test_date,
                report_date=report_date,
                accession_number=accession_number
            )
            pdf_model_instance.save()

            if re.fullmatch(r'\d{10}', accession_number):
                account_sid = settings.TWILIO_ACCOUNT_SID
                auth_token = settings.TWILIO_AUTH_TOKEN
                client = tw(account_sid, auth_token)
                #patient_name = 'Jangra'
                # media_url = 'media/uploads/xray_pdfs/RUCH1234_RUCHI%20JANGRA_IQ%20CITY%20ROAD.pdf'
                prefix = "media/uploads/xray_pdfs/"
                encoded_filename = pdf_file.name.replace(" ", "%20")
                media_url = prefix + encoded_filename
                message = client.messages.create(
                    content_sid='HXff6a8bf74ca42c765eefe580fb5b376b',
                    from_='MG228f0104ea3ddfc780cfcc1a0ca561d9',
                    to=f'whatsapp:+91{accession_number}',
                    content_variables=json.dumps({'1': patient_name, '2': media_url}),
                )
            
            return JsonResponse({'message': 'PDF successfully uploaded and processed.'})
        except Exception as e:
            print("Error processing PDF:", e)
            return JsonResponse({'error': 'Internal server error.'}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=400)


def get_csrf_token(request):
    csrf_token = {
        'csrf_token': get_token(request),

    }
    print("CSRF Token from headers:", csrf_token)
    return JsonResponse(csrf_token)


@user_type_required('xraycoordinator')
############################ To retrive data
# def xray_pdf_report(request):
#     pdfs = XrayReport.objects.all()
#     print([pdf.get_pdf_url() for pdf in pdfs])
#     return render(request, 'users/xray_pdf_report.html', {'pdfs': pdfs})

def xray_pdf_report(request):
    pdfs = XrayReport.objects.all().order_by('-report_date')

    # Collect unique dates and locations from the PDFs
    test_dates = set(pdf.test_date for pdf in pdfs)
    formatted_dates = [date.strftime('%Y-%m-%d') for date in test_dates]
    report_dates = set(pdf.report_date for pdf in pdfs)
    unique_locations = XLocation.objects.all()

    context = {
        'pdfs': pdfs,
        'Test_Date': sorted(formatted_dates),
        'Report_Date': sorted(report_dates),  # Ensure dates are sorted for dropdown
        'Location': unique_locations  # Ensure locations are sorted for dropdown
    }

    return render(request, 'users/xray_pdf_report.html', context)


@login_required
# def client_dashboard(request):
#     # Get the logged-in user's client object
#     user = request.user
#     try:
#         client = Client.objects.get(user=user)
#         user_location = client.location
#     except Client.DoesNotExist:
#         user_location = None

#     # Filter PDFs by the user's location
#     if user_location:
#         pdfs = XrayReport.objects.filter(location=user_location).order_by('-report_date')
#     else:
#         pdfs = XrayReport.objects.none()  # No PDFs if no location is set for the user

#     # Collect unique dates and locations from the PDFs
#     test_dates = set(pdf.test_date for pdf in pdfs)
#     formatted_dates = [date.strftime('%Y-%m-%d') for date in test_dates]
#     report_dates = set(pdf.report_date for pdf in pdfs)
#     unique_locations = XLocation.objects.all()

#     context = {
#         'pdfs': pdfs,
#         'Test_Date': sorted(formatted_dates),
#         'Report_Date': sorted(report_dates),  # Ensure dates are sorted for dropdown
#         'Location': unique_locations  # Ensure locations are sorted for dropdown
#     }

#     return render(request, 'users/client_dashboard.html', context)    


################################################################### Vitals PDF upload to portal #########################################################################

def handle_vitals_pdf_upload(pdf_file):
    your_pdf_model_instance = VitalsReport(pdf_file=pdf_file)
    your_pdf_model_instance.save()


def upload_vitals_pdf(request):
    if request.method == 'POST':
        try:
            print("Inside upload_vitals_pdf view")
            pdf_file = request.FILES.get('pdf')
            patient_id = request.POST.get('patientId')
            patient_name = request.POST.get('patientName')
            test_date_str = request.POST.get('testDate')
            report_date_str = request.POST.get('reportDate')

            if not pdf_file:
                return JsonResponse({'error': 'No PDF file provided.'}, status=400)

            # Specify the upload path and create a folder if it doesn't exist
            upload_path = os.path.join('uploads', 'vitals_pdfs')
            os.makedirs(upload_path, exist_ok=True)

            # Save the PDF file to the specified path
            pdf_file_path = os.path.join(upload_path, pdf_file.name)
            print("PDF file path:", pdf_file_path)

            with open(pdf_file_path, 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)

            # Convert report_date_str to a datetime object
            patient_id_str = request.POST.get('patientId')
            patient_id_str = patient_id_str.split(': ')[1]

            patient_name_str = request.POST.get('patientName')
            patient_name_str = patient_name_str.split(': ')[1]  # Remove the prefix "Test date: "
            # Extract the date string from the input
            test_date_str = request.POST.get('testDate')
            test_date_str = test_date_str.split(': ')[1]  # Remove the prefix "Test date: "
            # Extract the date string from the input
            report_date_str = request.POST.get('reportDate')
            report_date_str = report_date_str.split(': ')[1]  # Remove the prefix "Test date: "

            # Convert report_date_str to a datetime
            patient_id = patient_id_str
            patient_name = patient_name_str
            test_date = datetime.strptime(test_date_str, "%Y-%m-%d").date()
            report_date = datetime.strptime(report_date_str, "%Y-%m-%d").date()

            # Save the PDF file path and additional data to the database
            pdf_model_instance = VitalsReport(
                pdf_file=pdf_file_path,
                name=patient_name,
                patient_id=patient_id,
                test_date=test_date,
                report_date=report_date
            )
            pdf_model_instance.save()

            return JsonResponse({'message': 'PDF successfully uploaded and processed.'})
        except Exception as e:
            print("Error processing PDF:", e)
            return JsonResponse({'error': 'Internal server error.'}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=400)


def get_csrf_token(request):
    csrf_token = {
        'csrf_token': get_token(request),

    }
    print("CSRF Token from headers:", csrf_token)
    return JsonResponse(csrf_token)


@user_type_required('xraycoordinator')
def vitals_pdf_report(request):
    pdfs = VitalsReport.objects.all().order_by('-report_date')

    # Collect unique dates and locations from the PDFs
    test_dates = set(pdf.test_date for pdf in pdfs)
    formatted_dates = [date.strftime('%Y-%m-%d') for date in test_dates]
    report_dates = set(pdf.report_date for pdf in pdfs)

    context = {
        'pdfs': pdfs,
        'Test_Date': sorted(formatted_dates),
        'Report_Date': sorted(report_dates),  # Ensure dates are sorted for dropdown
    }

    return render(request, 'users/vitals_pdf_report.html', context)


###################################################################### END of vitals pdf upload #########################################################

################################################################### Audiometry PDF upload to portal #########################################################################

def handle_audiometry_pdf_upload(pdf_file):
    audiometry_pdf_instance = AudiometryReport(pdf_file=pdf_file)
    audiometry_pdf_instance.save()


def upload_audiometry_pdf(request):
    if request.method == 'POST':
        try:
            pdf_file = request.FILES.get('pdf')
            patient_id = request.POST.get('patientId')
            patient_name = request.POST.get('patientName')
            test_date_str = request.POST.get('testDate')
            report_date_str = request.POST.get('reportDate')

            print("Test date string:", test_date_str)
            print("Report date string:", report_date_str)

            if not pdf_file:
                return JsonResponse({'error': 'No PDF file provided.'}, status=400)

            # Convert date strings to datetime objects
            test_date = datetime.strptime(test_date_str, '%Y-%m-%d').date()
            report_date = datetime.strptime(report_date_str, '%Y-%m-%d').date()

            # Save the PDF file and additional data to the database
            pdf_model_instance = AudiometryReport(
                pdf_file=pdf_file,
                name=patient_name,
                patient_id=patient_id,
                test_date=test_date,
                report_date=report_date
            )
            pdf_model_instance.save()

            return JsonResponse({'message': 'PDF successfully uploaded and processed.'})
        except Exception as e:
            print("Error processing PDF:", e)
            return JsonResponse({'error': 'Internal server error.'}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=400)


@user_type_required('xraycoordinator')
def audiometry_pdf_report(request):
    pdfs = AudiometryReport.objects.all().order_by('-report_date')

    # Collect unique dates from the PDFs
    test_dates = set(pdf.test_date for pdf in pdfs)
    formatted_dates = [date.strftime('%Y-%m-%d') for date in test_dates]
    report_dates = set(pdf.report_date for pdf in pdfs)

    context = {
        'pdfs': pdfs,
        'Test_Date': sorted(formatted_dates),
        'Report_Date': sorted(report_dates),  # Ensure dates are sorted for dropdown
    }

    return render(request, 'users/audiometry_pdf_report.html', context)


###################################################################### END of Audiometry pdf upload #########################################################

################################################################### Optometry PDF upload to portal #########################################################################

def handle_optometry_pdf_upload(pdf_file):
    your_pdf_model_instance = OptometryReport(pdf_file=pdf_file)
    your_pdf_model_instance.save()


def upload_optometry_pdf(request):
    if request.method == 'POST':
        try:
            print("Inside upload_optometry_pdf view")
            pdf_file = request.FILES.get('pdf')
            patient_id = request.POST.get('patientId')
            patient_name = request.POST.get('patientName')
            test_date_str = request.POST.get('testDate')
            report_date_str = request.POST.get('reportDate')

            if not pdf_file:
                return JsonResponse({'error': 'No PDF file provided.'}, status=400)

            # Specify the upload path and create a folder if it doesn't exist
            upload_path = os.path.join('uploads', 'vitals_pdfs')
            os.makedirs(upload_path, exist_ok=True)

            # Save the PDF file to the specified path
            pdf_file_path = os.path.join(upload_path, pdf_file.name)
            print("PDF file path:", pdf_file_path)

            with open(pdf_file_path, 'wb+') as destination:
                for chunk in pdf_file.chunks():
                    destination.write(chunk)

            # Convert report_date_str to a datetime object
            patient_id_str = request.POST.get('patientId')
            patient_id_str = patient_id_str.split(': ')[1]

            patient_name_str = request.POST.get('patientName')
            patient_name_str = patient_name_str.split(': ')[1]  # Remove the prefix "Test date: "
            # Extract the date string from the input
            test_date_str = request.POST.get('testDate')
            test_date_str = test_date_str.split(': ')[1]  # Remove the prefix "Test date: "
            # Extract the date string from the input
            report_date_str = request.POST.get('reportDate')
            report_date_str = report_date_str.split(': ')[1]  # Remove the prefix "Test date: "

            # Convert report_date_str to a datetime
            patient_id = patient_id_str
            patient_name = patient_name_str
            test_date = datetime.strptime(test_date_str, "%Y-%m-%d").date()
            report_date = datetime.strptime(report_date_str, "%Y-%m-%d").date()

            # Save the PDF file path and additional data to the database
            pdf_model_instance = OptometryReport(
                pdf_file=pdf_file_path,
                name=patient_name,
                patient_id=patient_id,
                test_date=test_date,
                report_date=report_date
            )
            pdf_model_instance.save()

            return JsonResponse({'message': 'PDF successfully uploaded and processed.'})
        except Exception as e:
            print("Error processing PDF:", e)
            return JsonResponse({'error': 'Internal server error.'}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=400)


def get_csrf_token(request):
    csrf_token = {
        'csrf_token': get_token(request),

    }
    print("CSRF Token from headers:", csrf_token)
    return JsonResponse(csrf_token)


@user_type_required('xraycoordinator')
def optometry_pdf_report(request):
    pdfs = OptometryReport.objects.all().order_by('-report_date')

    # Collect unique dates and locations from the PDFs
    test_dates = set(pdf.test_date for pdf in pdfs)
    formatted_dates = [date.strftime('%Y-%m-%d') for date in test_dates]
    report_dates = set(pdf.report_date for pdf in pdfs)

    context = {
        'pdfs': pdfs,
        'Test_Date': sorted(formatted_dates),
        'Report_Date': sorted(report_dates),  # Ensure dates are sorted for dropdown
    }

    return render(request, 'users/optometry_pdf_report.html', context)


###################################################################### END of optometry pdf upload #########################################################
@user_type_required('xraycoordinator')
def get_excel(request):
    if request.method == 'GET':
        cities = XCity.objects.all()
        locations = XLocation.objects.all()
        test_dates = XrayReport.objects.filter().values_list('test_date', flat=True).distinct()
        formatted_dates = [date.strftime('%Y-%m-%d') for date in test_dates if date]
        return render(request, 'users/get_excel.html',
                      {'cities': cities, 'locations': locations, 'dates': formatted_dates})

    if request.method == 'POST':
        city_name = request.POST.get('city')
        location_name = request.POST.get('location')

        if city_name and location_name:
            location = XLocation.objects.get(name=location_name, city__name=city_name)
            pdf_files = XrayReport.objects.filter(location=location)
            patient_data_xray = []
            if pdf_files:
                for pdf_file in pdf_files:
                    with open(pdf_file.pdf_file.path, 'rb') as file:
                        pdf_reader = PyPDF2.PdfReader(file)
                        if len(pdf_reader.pages) > 0:
                            if len(pdf_reader.pages) == 1:
                                first_page = pdf_reader.pages[0]
                                first_page_text = first_page.extract_text()
                            else:
                                second_page = pdf_reader.pages[1]
                                second_page_text = second_page.extract_text()
                                first_page_text = second_page_text

                            try:
                                patient_id = str(first_page_text).split("Patient ID:")[1].split("Age:")[0].lower().strip()
                                patient_name = str(first_page_text).split("Name:")[1].split("Patient ID:")[0].lower().strip()
                                gender = str(first_page_text).split("Gender:")[1].split("Test date:")[0].strip().lower()
                                age = str(first_page_text).split("Age:")[1].split("YGender:")[0].strip().lower()
                                test_date = str(first_page_text).split("Test date:")[1].split("Report date:")[0].strip().lower()
                                report_date = str(first_page_text).split("Report date:")[1].split("X-RAY")[0].strip().lower()
                                findings = str(first_page_text).split("IMPRESSION:")[1].split("Dr.")[0].strip().lower()
                                print(patient_id, patient_name, gender, age, test_date, report_date, findings)
                                if "•" in findings:
                                    findings = findings.split("•")[1].split(".")[0]
                                else:
                                    findings = findings.split('.')[0]

                                if findings == 'No significant abnormality noted' or findings == 'No significant abnormality':
                                    findings = 'No significant abnormality seen'

                                patient_data_xray.append(
                                    (patient_id, patient_name, age, gender, test_date, report_date, findings.strip()))

                                print(patient_id, patient_name, age, gender, test_date, report_date, findings)
                            except IndexError as e:
                                print(f"Error processing file {pdf_file}: Invalid PDF Format")

                wb = Workbook()
                ws = wb.active
                ws.append(['Patient ID', 'Patient Name', 'Age', 'Gender', 'Test Date', 'Report Date', 'Findings'])

                # Define fill colors
                green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
                red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')

                # Write data to the Excel worksheet
                for data in patient_data_xray:
                    ws.append(data)
                    row_index = ws.max_row
                    findings_cell = ws.cell(row=row_index, column=ws.max_column)

                    if data[-1] == 'no significant abnormality seen':
                        findings_cell.fill = green_fill
                    else:
                        findings_cell.fill = red_fill

                # Save the workbook to a BytesIO buffer
                excel_buffer = BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

                # Create a response with the Excel file
                response = HttpResponse(
                    excel_buffer,
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                )
                response['Content-Disposition'] = 'attachment; filename=patient_data_xray.xlsx'

            else:
                context = {
                    'message': 'Currently, there are no PDF files in this location.'
                }
                response = render(request, 'users/get_excel.html', context)
        else:
            invalid_city = {'message': 'You do not have access to this city.'}
            response = render(request, 'users/get_excel.html', invalid_city)

        return response


def reject_patient_status(request, patient_id):
    try:
        patient = PatientDetails.objects.get(PatientId=patient_id)
        patient.status = True  # Update status to True to indicate rejection
        patient.save()
        return JsonResponse({'success': True, 'status': patient.status})
    except PatientDetails.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Patient not found'})


# def Auto_Sync(request):
#     root = Tk()
#     root.withdraw()
#
#     # Open directory dialog for selecting folder
#     folder_path = filedialog.askdirectory(title="Select folder containing DICOM files")
#     if folder_path:
#         dicom_instances = []
#         rejected_files = []
#
#         for root_dir, _, files in os.walk(folder_path):
#             for filename in files:
#                 if filename.lower().endswith('.dcm'):
#                     try:
#                         file_path = os.path.join(root_dir, filename)
#                         dicom_dataset = pydicom.dcmread(file_path)
#
#                         existing_instance = DICOMData.objects.filter(patient_id=str(dicom_dataset.PatientID)).first()
#                         if existing_instance:
#                             print(f"Skipping file {filename} - Duplicate data found.")
#                             rejected_files.append({'id': existing_instance.id, 'name': filename})
#                         else:
#                             print(f"Saving file {filename}")
#
#                             # Create DICOMData instance
#                             dicom_instance = DICOMData.objects.create(
#                                 patient_name=str(dicom_dataset.PatientName),
#                                 patient_id=str(dicom_dataset.PatientID),
#                                 age=str(dicom_dataset.PatientAge),
#                                 gender='Male' if dicom_dataset.PatientSex.upper() == 'M' else 'Female',
#                                 study_date=dicom_dataset.StudyDate if dicom_dataset.StudyDate else None,
#                                 study_description=str(dicom_dataset.StudyDescription),
#                                 notes=request.POST.get("note", "No Clinical History"),
#                                 body_part_examined=str(dicom_dataset.BodyPartExamined)
#                             )
#
#                             # Save the DICOM file
#                             dicom_file_obj = DICOMFile.objects.create(
#                                 dicom_data=dicom_instance,
#                                 dicom_file=filename
#                             )
#
#                             # Construct the file path for saving the DICOM file
#                             dicom_file_path = os.path.join('dicom_files', filename)
#
#                             # Write the DICOM file content to the file path
#                             with open(dicom_file_path, 'wb') as dicom_file:
#                                 dicom_file.write(dicom_dataset.PixelData)
#
#                             # Convert DICOM image to JPEG-compatible format
#                             pixel_data = dicom_dataset.pixel_array
#                             if dicom_dataset.BitsAllocated == 16:
#                                 pixel_data = pixel_data.astype('uint16')
#                                 pixel_data = pixel_data >> (dicom_dataset.BitsStored - 8)
#
#                             # Convert DICOM image to JPEG and save
#                             with BytesIO() as output:
#                                 Image.fromarray(pixel_data).convert('L').save(output, format='JPEG')
#
#                                 # Save the JPEG file with the correct DICOM instance
#                                 jpeg_file_name = f"{filename.split('.')[0]}.jpg"  # Assuming DICOM file name is unique
#                                 jpeg_file_content = output.getvalue()
#                                 jpeg_file = ContentFile(jpeg_file_content, name=jpeg_file_name)
#                                 jpeg_instance = JPEGFile.objects.create(dicom_data=dicom_instance, jpeg_file=jpeg_file)
#
#                             dicom_instances.append(dicom_instance)
#
#                     except Exception as e:
#                         print(f"Error saving file {filename}: {str(e)}")
#                         rejected_files.append({'id': None, 'name': filename})
#
#         # Update total uploaded X-ray count
#         total_cases, created = Total_Cases.objects.get_or_create(id=1, defaults={'total_uploaded_xray': 0})
#         total_cases.total_uploaded_xray += len(dicom_instances)
#         total_cases.save()
#
#         return JsonResponse({'message': 'Auto-sync process completed successfully.', 'rejected_files': rejected_files},
#                             status=200)
#     else:
#         return JsonResponse({'error': 'No folder path selected'}, status=400)

class GoogleDCM(View):
    def get(self, request):
        google_drive_data = GoogleDriveDCM()
        response = HttpResponse(google_drive_data)
        return response


def GoogleDriveDCM():
    # The file that contains the OAuth 2.0 credentials.
    CLIENT_SECRET_FILE = 'users/GoogleDriveAPI.json'

    # The name of the API and version of the API.
    API_NAME = 'drive'
    API_VERSION = 'v3'

    # The scopes that are required to access the API.
    SCOPES = ['https://www.googleapis.com/auth/drive']

    def create_service():
        # Create the credentials.
        creds = None
        if os.path.exists('token.pickle'):
            with open('token.pickle', 'rb') as token:
                creds = pickle.load(token)

        # If the credentials don't exist or are invalid, then create new ones.
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                # Create the flow object.
                flow = InstalledAppFlow.from_client_secrets_file(CLIENT_SECRET_FILE, SCOPES)

                # Run the flow to obtain the credentials.
                creds = flow.run_local_server(port=0)

                # Save the credentials for future use.
                with open('token.pickle', 'wb') as token:
                    pickle.dump(creds, token)
        # Create the service object.
        service = build(API_NAME, API_VERSION, credentials=creds)
        return service

    # folder_id = '1RjxYJcv4vbv1WFfcsUtCWm-qh0N3KRd0n'
    folder_id = '1_4MhRbCXyjtAFLnHbrP-KR_JimisCbwl'
    service = create_service()
    existing_patient_ids = set(DICOMData.objects.values_list('patient_id', flat=True))
    fetch_patient_data_from_folder1(service, folder_id, existing_patient_ids)


def fetch_patient_data_from_folder1(service, folder_id, existing_patient_ids):
    print('calling google drive...')
    main_folder = service.files().get(fileId=folder_id, fields='name').execute()
    main_folder_name = main_folder.get('name')

    subfolders_query = f"'{folder_id}' in parents and mimeType='application/vnd.google-apps.folder'"
    subfolders_response = service.files().list(q=subfolders_query, fields="files(id, name)").execute()
    subfolders = subfolders_response.get('files', [])

    dicom_instances = {}
    processed_patient_ids = set()

    for subfolder in subfolders:
        subfolder_id = subfolder.get('id')

        # Process files in the current subfolder
        files_query = f"'{subfolder_id}' in parents and trashed=false"
        files_response = service.files().list(q=files_query, fields="files(id, name, mimeType)").execute()
        files = files_response.get('files', [])

        excel_file_id = None

        for file in files:
            file_name = file.get('name')
            mime_type = file.get('mimeType')

            # Check if it's an Excel file
            if mime_type == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' and file_name.endswith('.xlsx'):
                excel_file_id = file.get('id')

                # Download the Excel file content
                request = service.files().get_media(fileId=excel_file_id)
                fh = BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()
                excel_data = pd.read_excel(fh)
                break

        if excel_file_id:
            break

    for subfolder in subfolders:
        subfolder_id = subfolder.get('id')
        subfolder_name = subfolder.get('name')

        try:
            location = XLocation.objects.get(name=subfolder_name)
            print(location)
        except Location.DoesNotExist:
            continue

        # Process files in the current subfolder
        files_query = f"'{subfolder_id}' in parents and trashed=false"
        files_response = service.files().list(q=files_query, fields="files(id, name, mimeType)").execute()
        files = files_response.get('files', [])

        for file in files:
            file_name = file.get('name')
            mime_type = file.get('mimeType')

            # Process only files with .dcm extension
            if mime_type == 'application/dicom' and file_name.endswith('.dcm'):
                file_id = file.get('id')

                # Download the DICOM file content
                request = service.files().get_media(fileId=file_id)
                fh = BytesIO()
                downloader = MediaIoBaseDownload(fh, request)
                done = False
                while done is False:
                    status, done = downloader.next_chunk()

                # Save the DICOM file locally in the 'dcm_files' folder
                dcm_file_path = os.path.join(r'C:\Users\admin\PycharmProjects\XRAY-REPORTING-master\authsysproject\dicom_files', file_name)

                with open(dcm_file_path, 'wb') as dcm_file:
                    dcm_file.write(fh.getvalue())

                dicom_data = pydicom.dcmread(dcm_file_path)

                # Extract patient information
                patient_id = dicom_data.get('PatientID')
                study_date_formatted = datetime.strptime(dicom_data.StudyDate, "%Y%m%d").strftime("%Y-%m-%d")

                if patient_id not in existing_patient_ids and patient_id not in processed_patient_ids:
                    # Create a new instance of DICOMData
                    dicom_instance = DICOMData.objects.create(
                        patient_id=patient_id,
                        patient_name=str(dicom_data.PatientName),
                        age=str(dicom_data.PatientAge),
                        gender='Male' if dicom_data.PatientSex.upper() == 'M' else 'Female',
                        location=location,
                        notes='No Clinical history.',
                        study_description=str(dicom_data.StudyDescription),
                        study_date=study_date_formatted
                    )

                    if excel_file_id:
                        patient_clinical_history = dict(zip(map(str, excel_data['patient_id']), excel_data['clinical_history']))
                        if dicom_instance.patient_id in patient_clinical_history.keys():
                            clinical_history = patient_clinical_history[patient_id]
                            if pd.isna(clinical_history) or (isinstance(clinical_history, float) and math.isnan(clinical_history)):
                                dicom_instance.notes = 'No Clinical History.'
                            else:
                                dicom_instance.notes = clinical_history
                            dicom_instance.save()

                    pixel_data = dicom_data.pixel_array
                    if dicom_data.BitsAllocated == 16:
                        pixel_data = pixel_data.astype('uint16')
                        pixel_data = pixel_data >> (dicom_data.BitsStored - 8)

                    # Convert DICOM image to JPEG
                    with BytesIO() as output:
                        Image.fromarray(pixel_data).convert('L').save(output, format='JPEG')

                        # Save the JPEG file with the correct DICOM instance
                        jpeg_file_name = f"{file_name.split('.')[0]}.jpg"
                        jpeg_file = ContentFile(output.getvalue(), name=jpeg_file_name)
                        JPEGFile.objects.create(dicom_data=dicom_instance, jpeg_file=jpeg_file)

                    print("Patient Saved:", dicom_instance.patient_name, dicom_instance.notes)
                    # Keep track of successfully processed instances
                    processed_patient_ids.add(patient_id)
                    dicom_instances[patient_id] = dicom_instance

    # Update total cases count
    total_cases, created = Total_Cases.objects.get_or_create(id=1, defaults={'total_uploaded_xray': 0})
    total_cases.total_uploaded_xray += len(dicom_instances)
    total_cases.save()


################################################ Auto reporting #####################################################
@user_type_required('campautomation')
def vitalslist(request):
    patients = vitalPatientDetails.objects.all()
    return render(request, 'users/vitalslist.html', {'patients': patients})

@user_type_required('campautomation')
def optometrylist(request):
    patients = PatientInfo.objects.all()
    return render(request, 'users/optometrylist.html', {'patients': patients})

# Everything related for the vaccination list: 

@user_type_required('campautomation')
def vaccinationlist(request):
    patients = vaccinationPatientDetails.objects.all()
    return render(request, 'users/vaccinationlist.html', {'patients': patients})

def delete_all_patients_for_vaccination(request):
    if request.method == 'POST':
        vaccinationPatientDetails.objects.all().delete()
        return redirect('vaccinationlist')
    return render(request, 'users/vaccinationlist.html')

def vaccinationpatientDetails(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = vaccinationPatientDetails.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId_icontains=query) | Q(PatientName_icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)


def add_patient_for_vaccination(request):
    if request.method == 'POST':
        try:
            # Retrieve patient details from the POST request
            patient_id = request.POST.get('PatientId')
            patient_name = request.POST.get('PatientName')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            hepatitis_e_batch_no = request.POST.get('Hepatitis_E_Batch_No')
            hepatitis_e_manufacturing_date = request.POST.get('Hepatitis_E_Manufacturing_Date')
            hepatitis_e_expiry_date = request.POST.get('Hepatitis_E_Expiry_Date')
            typhoid_batch_no = request.POST.get('Typhoid_Batch_No')
            typhoid_manufacturing_date = request.POST.get('Typhoid_Manufacturing_Date')
            typhoid_expiry_date = request.POST.get('Typhoid_Expiry_Date')
            # Create a new Patient object and save it to the database
            patient = vaccinationPatientDetails(
                PatientId=patient_id,
                PatientName=patient_name,
                age=age,
                gender=gender,
                Hepatitis_E_Batch_No=hepatitis_e_batch_no,
                Hepatitis_E_Manufacturing_Date=hepatitis_e_manufacturing_date,
                Hepatitis_E_Expiry_Date=hepatitis_e_expiry_date,
                Typhoid_Batch_No=typhoid_batch_no,
                Typhoid_Manufacturing_Date=typhoid_manufacturing_date,
                Typhoid_Expiry_Date=typhoid_expiry_date,

                # TestDate=test_date,
                # ReportDate=report_date,
                # height=height,
                # weight=weight,
                # blood=blood,
                # pulse=pulse,
            )
            patient.save()

            return redirect('vaccinationlist')
        except Exception as e:
            print("Error adding patient:", e)
            return JsonResponse({'error': 'Internal server error'}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
def uploadcsvforvaccination(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']

        # Adjust the field names according to your CSV file structure
        field_names = ['Timestamp','PatientName', 'PatientId', 'Age', 'Gender', 'Hepatitis_E_Batch_No', 'Hepatitis_E_Manufacturing_Date', 'Hepatitis_E_Expiry_Date', 'Typhoid_Batch_No',
                       'Typhoid_Manufacturing_Date', 'Typhoid_Expiry_Date']
    
        try:
            # Decode the CSV file data and split it into lines
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file, fieldnames=field_names)

            if reader.fieldnames == field_names:
                next(reader)

            # Initialize a list to store missing data logs
            missing_data_logs = []
            total_rows = 0
            saved_rows = 0

            # Store the CSV data in a variable
            csv_data = list(reader)

            # Check for missing data in each row
            for idx, row in enumerate(csv_data, start=1):
                total_rows += 1
                missing_fields = [field for field in field_names if not row.get(field)]
                if missing_fields:
                    # Append each missing data message separately for each row
                    error_message = f"Missing data for ID: {row.get('PatientId')} and Name: {row.get('PatientName')} in row {idx}: {', '.join(missing_fields)}"
                    missing_data_logs.append(error_message)
                    messages.error(request, error_message)
                else:
                    saved_rows += 1
                    # # Extract date and time from Timestamp
                     # Convert date strings to datetime objects


                      # Extract date and time from Timestamp
                    timestamp_str = row['Timestamp']
                    print(timestamp_str)


                    try:
                       # Try parsing with seconds included
                       timestamp_datetime = datetime.strptime(timestamp_str, '%d-%m-%Y %H:%M:%S')

                    except ValueError:
                        # If parsing with seconds fails, try without seconds
                        timestamp_datetime = datetime.strptime(timestamp_str, '%m/%d/%Y %H:%M:%S')

                    # Extract only the date part and format it as day/month/year
                    timestamp_date = timestamp_datetime.date()
                    timestamp_date_str = timestamp_date.strftime('%Y-%m-%d')
                    

                    # hepatitis_e_manufacturing_date = datetime.strptime(row['Hepatitis_E_Manufacturing_Date'], '%d-%m-%Y').date()
                    # hepatitis_e_expiry_date = datetime.strptime(row['Hepatitis_E_Expiry_Date'], '%d-%m-%Y').date()
                    # typhoid_manufacturing_date = datetime.strptime(row['Typhoid_Manufacturing_Date'], '%d-%m-%Y').date()
                    # typhoid_expiry_date = datetime.strptime(row['Typhoid_Expiry_Date'], '%d-%m-%Y').date()

                    # # Convert date to the desired format
                    # hepatitis_e_manufacturing_date_formatted = hepatitis_e_manufacturing_date.strftime('%Y-%m-%d')
                    # hepatitis_e_expiry_date_formatted = hepatitis_e_expiry_date.strftime('%Y-%m-%d')
                    # typhoid_manufacturing_date_formatted = typhoid_manufacturing_date.strftime('%Y-%m-%d')
                    # typhoid_expiry_date_formatted = typhoid_expiry_date.strftime('%Y-%m-%d')
                    

                    vaccinationPatientDetails.objects.create(
                        PatientId=row['PatientId'],
                        PatientName=row['PatientName'],
                        age=row['Age'],
                        gender=row['Gender'],
                        Hepatitis_E_Batch_No=row['Hepatitis_E_Batch_No'],
                        Hepatitis_E_Manufacturing_Date=row['Hepatitis_E_Manufacturing_Date'],
                        Hepatitis_E_Expiry_Date=row['Hepatitis_E_Expiry_Date'],
                        Typhoid_Batch_No=row['Typhoid_Batch_No'],
                        Typhoid_Manufacturing_Date=row['Typhoid_Manufacturing_Date'],
                        Typhoid_Expiry_Date=row['Typhoid_Expiry_Date'],
                        Date=timestamp_date_str,
                        # Hepatitis_E_Manufacturing_Date=hepatitis_e_manufacturing_date_formatted,
                        # Hepatitis_E_Expiry_Date=hepatitis_e_expiry_date_formatted,
                        # Typhoid_Batch_No=row['Typhoid_Batch_No'],
                        # Typhoid_Manufacturing_Date=typhoid_manufacturing_date_formatted,
                        # Typhoid_Expiry_Date=typhoid_expiry_date_formatted,
                    )

            if missing_data_logs:
                # Include total rows and saved rows in the error message
                error_message = f'\nTotal rows: {total_rows}, Saved rows: {saved_rows}'
                messages.error(request, error_message)
                return redirect('vaccinationlist')
            else:
                # Redirect to the vaccinationlist page after successful upload
                messages.success(request, 'CSV data uploaded successfully.')
                return redirect('vaccinationlist')

        except Exception as e:
            return HttpResponse(f'Error: {str(e)}')
    else:
        # return HttpResponse('Please upload a CSV file.')
        return render(request, 'users/uploadcsv.html')

# end for vaccination list.

# Everything Related to ECG List.

@user_type_required('campautomation')
def ecglist(request):
    patients = ecgPatientDetails.objects.all()
    return render(request, 'users/ecglist.html', {'patients': patients})

def delete_all_patients_for_ecg(request):
    if request.method == 'POST':
        ecgPatientDetails.objects.all().delete()
        return redirect('ecglist')
    return render(request, 'users/ecglist.html')

def ecgpatientDetails(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = ecgPatientDetails.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId_icontains=query) | Q(PatientName_icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)


def add_patient_for_ecg(request):
    if request.method == 'POST':
        try:
            # Retrieve patient details from the POST request
            patient_id = request.POST.get('PatientId')
            patient_name = request.POST.get('PatientName')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            testdate = request.POST.get('Test Date')
            reportdate = request.POST.get('Report Date')
            heartrate = request.POST.get('Heart Rate')
            findings = request.POST.get('Findings')
            # Create a new Patient object and save it to the database
            patient = ecgPatientDetails(
                PatientId=patient_id,
                PatientName=patient_name,
                age=age,
                gender=gender,
                testdate=testdate,
                reportdate=reportdate,
                heartrate=heartrate,
                findings=findings

                # TestDate=test_date,
                # ReportDate=report_date,
                # height=height,
                # weight=weight,
                # blood=blood,
                # pulse=pulse,
            )
            patient.save()

            return redirect('ecglist')
        except Exception as e:
            print("Error adding patient:", e)
            return JsonResponse({'error': 'Internal server error'}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
def uploadcsvforecg(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']

        # Adjust the field names according to your CSV file structure
        field_names = ['PatientId', 'PatientName', 'Age', 'Gender', 'TestDate', 'ReportDate', 'HeartRate', 'Findings']
    
        try:
            # Decode the CSV file data and split it into lines
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file, fieldnames=field_names)

            if reader.fieldnames == field_names:
                next(reader)

            # Initialize a list to store missing data logs
            missing_data_logs = []
            total_rows = 0
            saved_rows = 0

            # Store the CSV data in a variable
            csv_data = list(reader)

            # Check for missing data in each row
            for idx, row in enumerate(csv_data, start=1):
                total_rows += 1
                missing_fields = [field for field in field_names if not row.get(field)]
                if missing_fields:
                    # Append each missing data message separately for each row
                    error_message = f"Missing data for ID: {row.get('PatientId')} and Name: {row.get('PatientName')} in row {idx}: {', '.join(missing_fields)}"
                    missing_data_logs.append(error_message)
                    messages.error(request, error_message)
                else:
                    saved_rows += 1
                    # # Extract date and time from Timestamp
                     # Convert date strings to datetime objects
                    

                    ecgPatientDetails.objects.create(
                        PatientId=row['PatientId'],
                        PatientName=row['PatientName'],
                        age=row['Age'],
                        gender=row['Gender'],
                        testdate=row['TestDate'],
                        reportdate=row['ReportDate'],
                        heartrate=row['HeartRate'],
                        findings=row['Findings']
                    )

            if missing_data_logs:
                # Include total rows and saved rows in the error message
                error_message = f'\nTotal rows: {total_rows}, Saved rows: {saved_rows}'
                messages.error(request, error_message)
                return redirect('ecglist')
            else:
                # Redirect to the vaccinationlist page after successful upload
                messages.success(request, 'CSV data uploaded successfully.')
                return redirect('ecglist')

        except Exception as e:
            return HttpResponse(f'Error: {str(e)}')
    else:
        # return HttpResponse('Please upload a CSV file.')
        return render(request, 'users/uploadcsv.html')
    


# End of ECG list. - Himanshu.

# Everything Related to xray List.

@user_type_required('campautomation')
def xraylist(request):
    patients = xrayPatientDetails.objects.all()
    return render(request, 'users/xraylist.html', {'patients': patients})

def delete_all_patients_for_xray(request):
    if request.method == 'POST':
        xrayPatientDetails.objects.all().delete()
        return redirect('xraylist')
    return render(request, 'users/xraylist.html')

def xraypatientDetails(request):
    if request.method == 'GET':
        query = request.GET.get('query', None)
        patients = xrayPatientDetails.objects.all()
        if query is not None:
            patients = patients.filter(Q(PatientId_icontains=query) | Q(PatientName_icontains=query))
        # response = {"patients": patients}
        response = serialize("json", patients)
        response = json.loads(response)
        return JsonResponse(status=200, data=response, safe=False)


def add_patient_for_xray(request):
    if request.method == 'POST':
        print(request.POST)
        
        try:
            # Retrieve patient details from the POST request
            patient_id = request.POST.get('PatientId')
            patient_name = request.POST.get('PatientName')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            testdate = request.POST.get('Test Date')
            reportdate = request.POST.get('Report Date')
            impressions = request.POST.get('Impression')
            findings = request.POST.get('Findings')
            jpeg_file = request.FILES.get('xray_image')
            # Create a new Patient object and save it to the database
            patient = xrayPatientDetails(
                PatientId=patient_id,
                PatientName=patient_name,
                age=age,
                gender=gender,
                testdate=testdate,
                reportdate=reportdate,
                impressions=impressions,
                findings=findings,
                jpeg_file=jpeg_file
            )
            patient.save()

            return redirect('xraylist')
        except Exception as e:
            print("Error adding patient:", e)
            return JsonResponse({'error': 'Internal server error'}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)
    
def uploadcsvforxray(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']

        # Adjust the field names according to your CSV file structure
        field_names = ['PatientId', 'PatientName', 'Age', 'Gender', 'TestDate', 'ReportDate', 'Findings', 'Impression']
    
        try:
            # Decode the CSV file data and split it into lines
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file, fieldnames=field_names)

            if reader.fieldnames == field_names:
                next(reader)

            # Initialize a list to store missing data logs
            missing_data_logs = []
            total_rows = 0
            saved_rows = 0

            # Store the CSV data in a variable
            csv_data = list(reader)

            # Check for missing data in each row
            for idx, row in enumerate(csv_data, start=1):
                total_rows += 1
                missing_fields = [field for field in field_names if not row.get(field)]
                if missing_fields:
                    # Append each missing data message separately for each row
                    error_message = f"Missing data for ID: {row.get('PatientId')} and Name: {row.get('PatientName')} in row {idx}: {', '.join(missing_fields)}"
                    missing_data_logs.append(error_message)
                    messages.error(request, error_message)
                else:
                    # Search for the corresponding image in the JPEGFile model
                    dicom_data = DICOMData.objects.filter(patient_id=row['PatientId']).first()
                    jpeg_file_instance = None
                    if dicom_data:
                        jpeg_file_instance = JPEGFile.objects.filter(dicom_data=dicom_data).first()

                    # This is the log to check whether the image is getting fetched or not . - Himanshu.
                    # if jpeg_file_instance:
                    #     # Print the image name for debugging
                    #     print(f"Image found for Patient ID {row['PatientId']}: {jpeg_file_instance.jpeg_file.name}")
                    # else:
                    #     print(f"No image found for Patient ID {row['PatientId']}")

                    # Save the xrayPatientDetails object with the associated image if found
                    saved_rows += 1
                    xrayPatientDetails.objects.create(
                        PatientId=row['PatientId'],
                        PatientName=row['PatientName'],
                        age=row['Age'],
                        gender=row['Gender'],
                        testdate=row['TestDate'],
                        reportdate=row['ReportDate'],
                        impressions=row['Impression'],
                        findings=row['Findings'],
                        jpeg_file=jpeg_file_instance.jpeg_file if jpeg_file_instance else None
                    )

            if missing_data_logs:
                # Include total rows and saved rows in the error message
                error_message = f'\nTotal rows: {total_rows}, Saved rows: {saved_rows}'
                messages.error(request, error_message)
                return redirect('xraylist')
            else:
                # Redirect to the xraylist page after successful upload
                messages.success(request, 'CSV data uploaded successfully.')
                return redirect('xraylist')

        except Exception as e:
            return HttpResponse(f'Error: {str(e)}')
    else:
        # return HttpResponse('Please upload a CSV file.')
        return render(request, 'users/uploadcsv.html')
    


# End of Xray list. - Himanshu.

def delete_all_patients_opto(request):
    if request.method == 'POST':
        PatientInfo.objects.all().delete()
        return redirect('optometrylist')
    return render(request, 'users/optometrylist.html')


def delete_all_patients(request):
    if request.method == 'POST':
        vitalPatientDetails.objects.all().delete()
        return redirect('vitalslist')
    return render(request, 'users/vitalslist.html')



def add_patient(request):
    if request.method == 'POST':
        try:
            # Retrieve patient details from the POST request
            patient_name = request.POST.get('PatientName')
            patient_id = request.POST.get('PatientId')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            test_date = request.POST.get('TestDate')
            report_date = request.POST.get('ReportDate')
            height = request.POST.get('height')
            weight = request.POST.get('weight')
            blood = request.POST.get('blood')
            pulse = request.POST.get('pulse')
            # Create a new Patient object and save it to the database
            patient = vitalPatientDetails(
                PatientId=patient_id,
                PatientName=patient_name,
                age=age,
                gender=gender,
                TestDate=test_date,
                ReportDate=report_date,
                height=height,
                weight=weight,
                blood=blood,
                pulse=pulse,
            )
            patient.save()

            return redirect('vitalslist')
        except Exception as e:
            print("Error adding patient:", e)
            return JsonResponse({'error': 'Internal server error'}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


def add_patient_opto(request):
    if request.method == 'POST':
        try:
            # Retrieve patient details from the POST request
            patient_name = request.POST.get('PatientName')
            patient_id = request.POST.get('PatientId')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            test_date = request.POST.get('TestDate')
            report_date = request.POST.get('ReportDate')
            FarVisionRight = request.POST.get('FarVisionRight')
            FarVisionLeft = request.POST.get('FarVisionLeft')
            NearVisionRight = request.POST.get('NearVisionRight')
            NearVisionLeft = request.POST.get('NearVisionLeft')
            ColorBlindness = request.POST.get('ColorBlindness')
            # Create a new Patient object and save it to the database
            patient = PatientInfo(
                PatientId=patient_id,
                PatientName=patient_name,
                age=age,
                gender=gender,
                TestDate=test_date,
                ReportDate=report_date,
                FarVisionRight=FarVisionRight,
                FarVisionLeft=FarVisionLeft,
                NearVisionRight=NearVisionRight,
                NearVisionLeft=NearVisionLeft,
                ColorBlindness=ColorBlindness,
            )
            patient.save()

            return redirect('optometrylist')
        except Exception as e:
            print("Error adding patient:", e)
            return JsonResponse({'error': 'Internal server error'}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


def uploadcsv(request):
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']

        # Adjust the field names according to your CSV file structure
        field_names = ['PatientId', 'PatientName', 'age', 'gender', 'TestDate', 'ReportDate', 'FarVisionRight',
                       'FarVisionLeft', 'NearVisionRight', 'NearVisionLeft',
                       'SphericalRight', 'CylindricalRight', 'AxisRight', 'AddRight', 
                       'SphericalLeft', 'CylindricalLeft', 'AxisLeft', 'AddLeft', 'ColorBlindness']

        try:
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file, fieldnames=field_names)
            if reader.fieldnames == field_names:
                next(reader)

            # Initialize a list to store missing data logs
            missing_data_logs = []
            total_rows = 0
            saved_rows = 0

            # Store the CSV data in a variable
            csv_data = list(reader)

            # Check for missing data in each row
            for idx, row in enumerate(csv_data, start=1):
                total_rows += 1
                missing_fields = [field for field in field_names if not row.get(field)]
                if missing_fields:
                    # Append each missing data message separately for each row
                    error_message = f"Missing data for ID: {row.get('PatientId')} and Name: {row.get('PatientName')} in row {idx}: {', '.join(missing_fields)}"
                    missing_data_logs.append(error_message)
                    messages.error(request, error_message)
                else:
                    saved_rows += 1
                    # Convert date strings to datetime objects
                    test_date = datetime.strptime(row['TestDate'], '%d-%m-%Y').date()
                    report_date = datetime.strptime(row['ReportDate'], '%d-%m-%Y').date()

                    # Convert datetime objects back to strings in the desired format
                    test_date_formatted = test_date.strftime('%Y-%m-%d')
                    report_date_formatted = report_date.strftime('%Y-%m-%d')

                    # Standardize date format and replace month names with digits
                    far_vision_left = row['FarVisionLeft'].replace('Jan', '1').replace('Feb', '2').replace('Mar',
                                                                                                           '3').replace(
                        'Apr', '4').replace('May', '5').replace('Jun', '6').replace('Jul', '7').replace('Aug',
                                                                                                        '8').replace(
                        'Sep', '9').replace('Oct', '10').replace('Nov', '11').replace('Dec', '12').replace('-', '/')
                    far_vision_right = row['FarVisionRight'].replace('Jan', '1').replace('Feb', '2').replace('Mar',
                                                                                                             '3').replace(
                        'Apr', '4').replace('May', '5').replace('Jun', '6').replace('Jul', '7').replace('Aug',
                                                                                                        '8').replace(
                        'Sep', '9').replace('Oct', '10').replace('Nov', '11').replace('Dec', '12').replace('-', '/')

                    # Remove leading zeros
                    far_vision_left = far_vision_left.lstrip('0')
                    far_vision_right = far_vision_right.lstrip('0')

                    PatientInfo.objects.create(
                        PatientId=row['PatientId'],
                        PatientName=row['PatientName'],
                        age=row['age'],
                        gender=row['gender'],
                        TestDate=test_date_formatted,
                        ReportDate=report_date_formatted,
                        FarVisionRight=far_vision_right,
                        FarVisionLeft=far_vision_left,
                        NearVisionRight=row['NearVisionRight'],
                        NearVisionLeft=row['NearVisionLeft'],
                        ColorBlindness=row['ColorBlindness'],
                        SphericalRight=row['SphericalRight'],
                        CylindricalRight=row['CylindricalRight'],
                        AxisRight=row['AxisRight'],
                        AddRight=row['AddRight'],
                        SphericalLeft=row['SphericalLeft'],
                        CylindricalLeft=row['CylindricalLeft'],
                        AxisLeft=row['AxisLeft'],
                        AddLeft=row['AddLeft']
                    )

            if missing_data_logs:
                # Include total rows and saved rows in the error message
                error_message = f'\nTotal rows: {total_rows}, Saved rows: {saved_rows}'
                messages.error(request, error_message)
                return redirect('optometrylist')
            else:
                # Redirect to the optometrylist page after successful upload
                messages.success(request, 'CSV data uploaded successfully.')
                return redirect('optometrylist')

        except Exception as e:
            return HttpResponse(f'Error: {str(e)}')

    else:
        # return HttpResponse('Please upload a CSV file.')
        return render(request, 'users/uploadcsv.html')


def add_patient_audio(request):
    if request.method == 'POST':
        try:
            # Retrieve patient details from the POST request
            patient_name = request.POST.get('PatientName')
            patient_id = request.POST.get('PatientId')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            test_date = request.POST.get('TestDate')
            report_date = request.POST.get('ReportDate')
            rightEarDB = request.POST.get('rightEarDB')
            leftEarDB = request.POST.get('leftEarDB')
            rightEarBoneDB = request.POST.get('rightEarBoneDB')
            leftEarBoneDB = request.POST.get('leftEarBoneDB')
            rightEarLevel = request.POST.get('rightEarLevel')
            leftEarLevel = request.POST.get('leftEarLevel')
            # Create a new Patient object and save it to the database
            patient = audioPatientDetails(
                PatientId=patient_id,
                PatientName=patient_name,
                age=age,
                gender=gender,
                TestDate=test_date,
                ReportDate=report_date,
                rightEarDB=rightEarDB,
                leftEarDB=leftEarDB,
                rightEarBoneDB=rightEarBoneDB,
                leftEarBoneDB=leftEarBoneDB,
                rightEarLevel=rightEarLevel,
                leftEarLevel=leftEarLevel,
            )
            patient.save()

            return redirect('audiometry')
        except Exception as e:
            print("Error adding patient:", e)
            return JsonResponse({'error': 'Internal server error'}, status=500)
    else:
        return JsonResponse({'error': 'Invalid request method'}, status=400)


def delete_all_patients_audio(request):
    if request.method == 'POST':
        audioPatientDetails.objects.all().delete()
        return redirect('audiometry')
    return render(request, 'users/audiometry.html')


def uploadcsvforvital(request):
    if request.method == 'POST' and request.FILES['csv_file']:
        csv_file = request.FILES['csv_file']

        # Adjust the field names according to your CSV file structure
        field_names = ['PatientId', 'PatientName', 'Age', 'Gender', 'TestDate', 'ReportDate', 'Height', 'Weight',
                       'Blood', 'Pulse']

        try:
            # Decode the CSV file data and split it into lines
            decoded_file = csv_file.read().decode('utf-8').splitlines()
            reader = csv.DictReader(decoded_file, fieldnames=field_names)
            if reader.fieldnames == field_names:
                next(reader)

            # Initialize a list to store missing data logs
            missing_data_logs = []
            total_rows = 0
            saved_rows = 0

            # Store the CSV data in a variable
            csv_data = list(reader)

            # Check for missing data in each row
            for idx, row in enumerate(csv_data, start=1):
                total_rows += 1
                missing_fields = [field for field in field_names if not row.get(field)]
                if missing_fields:
                    # Append each missing data message separately for each row
                    error_message = f"Missing data for ID: {row.get('PatientId')} and Name: {row.get('PatientName')} in row {idx}: {', '.join(missing_fields)}"
                    missing_data_logs.append(error_message)
                    messages.error(request, error_message)
                else:
                    saved_rows += 1
                    # Extract date and time from Timestamp
                    test_date = datetime.strptime(row['TestDate'], '%d-%m-%Y').date()
                    report_date = datetime.strptime(row['ReportDate'], '%d-%m-%Y').date()

                    # Convert date to the desired format
                    test_date_formatted = test_date.strftime('%Y-%m-%d')
                    report_date_formatted = report_date.strftime('%Y-%m-%d')

                    vitalPatientDetails.objects.create(
                        PatientId=row['PatientId'],
                        PatientName=row['PatientName'],
                        age=row['Age'],
                        gender=row['Gender'],
                        TestDate=test_date_formatted,
                        ReportDate=report_date_formatted,
                        height=row['Height'],
                        weight=row['Weight'],
                        blood=row['Blood'],
                        pulse=row['Pulse'],
                    )

            if missing_data_logs:
                # Include total rows and saved rows in the error message
                error_message = f'\nTotal rows: {total_rows}, Saved rows: {saved_rows}'
                messages.error(request, error_message)
                return redirect('vitalslist')
            else:
                # Redirect to the optometrylist page after successful upload
                messages.success(request, 'CSV data uploaded successfully.')
                return redirect('vitalslist')

        except Exception as e:
            return HttpResponse(f'Error: {str(e)}')
    else:
        # return HttpResponse('Please upload a CSV file.')
        return render(request, 'users/uploadcsv.html')


def patient_id_list(request):
    patient_ids = DICOMData.objects.values_list('patient_id', flat=True)
    return JsonResponse(list(patient_ids), safe=False)

@user_type_required('xraycoordinator')
def ReportingStatus(request):
    selected_date = request.GET.get('selected_date')
    total_count_data = []

    for client in XClient.objects.all():
        total_uploaded = 0
        total_assigned = 0
        total_reported = 0

        for city in client.xcity_set.all():
            for location in city.xlocation_set.all():
                dicom_data = DICOMData.objects.filter(location=location)
                if selected_date:
                    dicom_data = dicom_data.filter(study_date=selected_date)

                total_uploaded += dicom_data.count()
                total_assigned += dicom_data.filter(radiologist__isnull=False, isDone=False).values('patient_id').distinct().count()
                total_reported += dicom_data.filter(isDone=True).count()

        pending_reports = total_uploaded - total_reported

        total_count_data.append({
            'client_name': client.name,
            'total_count': total_uploaded,
            'total_assigned': total_assigned,
            'total_reported': total_reported,
            'pending_reports': pending_reports
        })

    for set_count_obj in SetCount.objects.all():
        found = False
        for item in total_count_data:
            if set_count_obj.client.name == item['client_name']:
                item['total_proposal'] = set_count_obj.TotalProposalbyClient
                item['total_done'] = set_count_obj.TotalCasesDone
                found = True
                break
        if not found:
            total_count_data.append({
                'client_name': set_count_obj.client.name,
                'total_proposal': set_count_obj.TotalProposalbyClient,
                'total_done': set_count_obj.TotalCasesDone
            })

    context = {
        'total_count_data': total_count_data
    }
    return render(request, 'users/reporting_status.html', context)

@user_type_required('xraycoordinator')
def SetTarget(request):
    if request.method == 'GET':
        clients = XClient.objects.all()
        return render(request, 'users/set_count.html', {'clients': clients})
    elif request.method == 'POST':
        client_id = request.POST.get('clientName')
        date = request.POST.get('date')
        total_proposal = request.POST.get('totalProposal')
        total_done_by_tech = request.POST.get('TotalCasesDone')

        client = XClient.objects.get(pk=client_id)
        existing_set_count = SetCount.objects.filter(client=client, date_field=date).first()
        if existing_set_count:
            existing_set_count.TotalProposalbyClient = total_proposal
            existing_set_count.TotalCasesDone = total_done_by_tech
            existing_set_count.save()
        else:
            set_count = SetCount.objects.create(client=client, TotalProposalbyClient=total_proposal, TotalCasesDone=total_done_by_tech, date_field=date)
            set_count.save()

        return redirect('reporting_status')


@user_type_required('ecgcoordinator')
def ECGReportingStatus(request):
    selected_date = request.GET.get('selected_date')
    total_count_data = []

    for client in Client.objects.all():
        total_uploaded = 0
        total_assigned = 0
        total_reported = 0

        for city in client.city_set.all():
            for location in city.location_set.all():
                patient_data = PatientDetails.objects.filter(location=location)
                print('patient data', patient_data)
                if selected_date:
                    patient_data = patient_data.filter(TestDate=selected_date)

                total_uploaded += patient_data.count()
                total_assigned += patient_data.filter(cardiologist__isnull=False, isDone=False).values('PatientId').distinct().count()
                total_reported += patient_data.filter(isDone=True).count()

        pending_reports = total_uploaded - total_reported

        total_count_data.append({
            'client_name': client.name,
            'total_count': total_uploaded,
            'total_assigned': total_assigned,
            'total_reported': total_reported,
            'pending_reports': pending_reports
        })

    for set_count_obj in ECGSetCount.objects.all():
        found = False
        for item in total_count_data:
            if set_count_obj.client.name == item['client_name']:
                item['total_proposal'] = set_count_obj.TotalProposalbyClient
                item['total_done'] = set_count_obj.TotalCasesDone
                found = True
                break
        if not found:
            total_count_data.append({
                'client_name': set_count_obj.client.name,
                'total_proposal': set_count_obj.TotalProposalbyClient
            })

    context = {
        'total_count_data': total_count_data
    }
    return render(request, 'users/ecg_reporting_status.html', context)


def ECGSetTarget(request):
    if request.method == 'GET':
        clients = Client.objects.all()
        return render(request, 'users/ecg_set_count.html', {'clients': clients})
    elif request.method == 'POST':
        client_id = request.POST.get('clientName')
        date = request.POST.get('date')
        total_proposal = request.POST.get('totalProposalECG')
        total_done_by_tech = request.POST.get('TotalCasesDoneECG')

        client = Client.objects.get(pk=client_id)
        existing_set_count = ECGSetCount.objects.filter(client=client, date_field=date).first()
        if existing_set_count:
            existing_set_count.TotalProposalbyClient = total_proposal
            existing_set_count.TotalCasesDone = total_done_by_tech
            existing_set_count.save()
        else:
            set_count = ECGSetCount.objects.create(client=client, TotalProposalbyClient=total_proposal, TotalCasesDone=total_done_by_tech, date_field=date)
            set_count.save()

        return redirect('ecg_reporting_status')

##################################################################################################################
# Everything related to uploading the ecg files from the ecg coordinator dashboard. - Himanshu.
def extract_patient_id(text):
    try:
        if "Id :" in str(text):
            id = str(text).split("Id :")[1].split(" ")[1].split("\n")[0].strip().lower()
            # print("This is the first fetched id :", id)
            if id == '':
                fetched_id = str(text).split("Id :")[1].split("Name :")[0].strip().lower()
                id = fetched_id.replace(" ", "")
                if id == '':
                    id = str(text).split("Comments")[1].split("HR")[0].strip()
                    # print("comments", id)
                    if id == '':
                        print("Id is not mentioned in the file. ")
                        print("Id :" , id)
        if "Id:" in str(text):
            id = str(text).split("Id:")[1].split(" ")[1].split("\n")[0].strip().lower()
            # print("This is the first fetched id :", id)
            if id == '':
                fetched_id = str(text).split("Id:")[1].split("Name :")[0].strip().lower()
                id = fetched_id.replace(" ", "")
                if id == '':
                    id = str(text).split("Comments")[1].split("HR")[0].strip()
                    # print("comments", id)
                    # Adding the case to remove the extra space issue.
                    if id == '':
                        id = str(text).split("Id:")[1].split("Name:")[0].strip()
                        if id == '':
                            print("Id is not mentioned in the file. ")
                            print("Id :" , id)
            
        return id
    except IndexError:
        return 'Missing'
    # try:
        # id = str(text).split("Id")[1].split("\n")[0]
        # print(id)
        # if ":" in id:
        #     return id.split(":")[1].strip()
        # else:
        #     return id.strip()
    # except IndexError:
        # Handle cases where the expected format is not found
        # if text.count('Comments') > 1:
        #     return str(text).split("Comments\nComments")[1].split("HR")[0].split('\n')[1].split('\n')[0]
        # else:
        #     return str(text).split("Comments")[1].split("HR")[0].strip()
        

def extract_patient_name(text):
    try:
        return str(text).split("Name")[1].split("\n")[0].split(":")[1].strip()
    except IndexError:
        return 'None'

def extract_patient_age(text):
    try:
        return str(text).split("Age")[1].split("\n")[0].split(":")[1].strip()
    except IndexError:
        return '0'  # Default age if not found

def extract_patient_gender(text):
    try:
        return str(text).split("Gender")[1].split("\n")[0].split(":")[1].strip()
    except IndexError:
        return 'Missing'

def extract_heart_rate(text):
    try:
        hr = str(text).split("HR:")[1].split("/")[0].strip()
        print("This is the Heart Rate :", hr)
        # if hr == '':
        #     hr = str(text).split("HR:")[1].split("/")[0].strip()
        #     print("This is updated Heart Rate :", hr)
        return hr
    except IndexError:
        return '0'  # Default heart rate if not found

def extract_pr_interval(text):
    try:
        return str(text).split("PR:")[1].split("QRS:")[0].split("ms")[0].strip()
    except IndexError:
        return '0'  # Default PR interval if not found

def extract_report_time(text):
    try:
        return str(text).split("Acquired on:")[1][12:17].strip()
    except IndexError:
        return '00:00'  # Default time if not found

def extract_date(text):
    try:
        raw_date = str(text).split("Acquired on:")[1][0:11].strip()
        if isinstance(raw_date, str):
            return datetime.strptime(raw_date, '%Y-%m-%d').date()
        else:
            return raw_date  # If raw_date is already a datetime.date, return it as is
    except (IndexError, ValueError):
        return datetime.now().date()  # Default to current date if not found

# To fix the duplicate extraction of some ecg graph pdf's.
def deduplicate_text(text):
    lines = text.split('\n')
    unique_lines = list(dict.fromkeys(lines))
    return '\n'.join(unique_lines)

# I am adding this function so that it can solve the issue of space after each character.
def clean_page_data(first_page_text):
    # Split the input data by lines
    lines = first_page_text.split('\n')
    
    # Initialize a list to hold cleaned lines
    cleaned_lines = []
    
    # Iterate through each line
    for line in lines:
        # Remove spaces after each character by replacing ' ' with '' and then joining characters
        cleaned_line = ''.join(line.split())
        
        # Append the cleaned line to the list
        cleaned_lines.append(cleaned_line)
    
    # Join the cleaned lines with newline characters for final output
    return '\n'.join(cleaned_lines)




def upload_ecg(request):
    success_message = ''
    success_details = []
    rejected_message = ''
    rejected_details = []

    # This is not needed now .------
    if len(rejected_details) != 0:
        rejected_details.clear() 
    # -------------------------------

    
    if request.method == 'POST':
        form = ECGUploadForm(request.POST, request.FILES)
        # This is logs for debugging. - Himanshu.
        # print("Submitted data:", request.POST)  # Print POST data
        # print("Submitted files:", request.FILES)  # Print uploaded files
        if form.is_valid():
            print("Form is valid.")
            ecg_files = form.cleaned_data['ecg_file'] # This will be a list of files.
            location = form.cleaned_data['location']
            print("Selected location:", location)

            for ecg_file in ecg_files:
                try:
                    pdf_bytes = ecg_file.read()
                    # print(pdf_bytes)
                except Exception as e:
                    print(f"Error reading ECG file: {str(e)}")
                    rejected_details.append({'id': None, 'name': ecg_file.name})
                    continue

                # pdf_bytes = ecg_file.read()
                pdf_reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))

                for page_number, page in enumerate(pdf_reader.pages):
                    first_page_text = page.extract_text()
                    # Checking if duplicate data is present or not.
                    first_page_text = deduplicate_text(first_page_text)
                    # Passing the deduplicated data.
                    print("This is the extracted first page text", first_page_text)

                    # Fixing the page text if it contains space after each character.
                    # Adding a flag.
                    extraSpace = False
                    if "A c q u i r e d  o n :" in first_page_text:
                        first_page_text = clean_page_data(first_page_text)
                        # Printing the cleaned text.
                        print("This is the cleaned text after removing the extra space issue :")
                        print(first_page_text)
                        # Now i am extracting the id separately in this condition.
                        patient_id = str(first_page_text).split("Id:")[1].split("Name:")[0].strip()
                        # Changing the flag to true.
                        extraSpace = True

                        


                    # Extract patient data using custom extraction functions
                    if extraSpace == False:
                        patient_id = extract_patient_id(first_page_text)
                        print("This is the patient id :", patient_id)
                        # Changing the flag back to False.
                        extraSpace = True

                    # Rejecting the file if the id is not present there.
                    # if patient_id == '':
                    #     print(f"Skipping file {ecg_file.name} - Id is not present in the uploaded file.")
                    #     rejected_details.append({'id': patient_id, 'name': ecg_file.name})

                    # Check if the patient ID is valid
                    if not patient_id:
                        print(f"Skipping file {ecg_file.name} - Id is not present in the uploaded file.")
                        rejected_details.append({'id': patient_id, 'name': ecg_file.name})
                        break  # Break out of the inner loop to skip the current file
                    
                    if PatientDetails.objects.filter(PatientId=patient_id).exists():
                        print(f"Skipping file {ecg_file.name} - Duplicate data found.")
                        rejected_details.append({'id': patient_id, 'name': ecg_file.name})

                    if not PatientDetails.objects.filter(PatientId=patient_id).exists():
                        patient_name = extract_patient_name(first_page_text)
                        patient_age = extract_patient_age(first_page_text)
                        patient_gender = extract_patient_gender(first_page_text)
                        heart_rate = extract_heart_rate(first_page_text)
                        pr_interval = extract_pr_interval(first_page_text)
                        report_time = extract_report_time(first_page_text)
                        formatted_date = extract_date(first_page_text)

                        # Create or get Date instance
                        date, created = Date.objects.get_or_create(date_field=formatted_date, location_id=location.id)

                        # Save PatientDetails
                        patient = PatientDetails(
                            PatientId=patient_id,
                            PatientName=patient_name,
                            age=patient_age,
                            gender=patient_gender,
                            HeartRate=heart_rate,
                            PRInterval=pr_interval,
                            TestDate=formatted_date,
                            ReportDate=formatted_date,
                            date=date,
                            location=location
                        )
                        patient.save()

                        # Convert PDF page to image and save to patient.image
                        doc = fitz.open(stream=pdf_bytes, filetype='pdf')
                        page = doc.load_page(page_number)
                        image_bytes = page.get_pixmap().tobytes()
                        image_buffer = io.BytesIO(image_bytes)
                        image_file = ContentFile(image_buffer.getvalue(), name=f"{patient_id}_{patient_name}.jpg")
                        patient.image.save(image_file.name, image_file)

                        # Save PDF file to patient.reportimage
                        reportimage_file = ContentFile(pdf_bytes, name=f"{patient_id}_{patient_name}.pdf")
                        patient.reportimage.save(reportimage_file.name, reportimage_file)
                        patient.save()
                        print("Patient saved successfully.")
                        success_details.append({'id': patient_id, 'name': ecg_file.name})
            
            # Updating the total ecg cases.

            # Retrieving total_cases
            total_cases, created = Total_Cases.objects.get_or_create(id=1, defaults={'total_uploaded_ecg': 0})

            # Update total cases count
            total_cases.total_uploaded_ecg += len(success_details)
            total_cases.save()

            if rejected_details:
                    rejected_message = f"{len(rejected_details)} files were rejected. Please check and try again."
                    rejected_details = [{'id': item['id'], 'name': item['name']} for item in rejected_details]

            if success_details:
                success_message = f"{len(success_details)} Images uploaded successfully."
                success_details = [{'id': item['id'], 'name': item['name']} for item in success_details]
                
            #########################################################################################
            # Passing all the data of allocation page.
            # Fetch and order patients
            patients = PatientDetails.objects.all().order_by('-TestDate')
            
            # Total counts for statistics
            total_current_uploaded = PatientDetails.objects.all().count()
            total_uploaded_ecg = Total_Cases.objects.values_list('total_uploaded_ecg', flat=True).first()
            total_reported_ecg = Total_Cases.objects.values_list('total_reported_ecg', flat=True).first()
            total_nonreported_ecg = Total_Cases.objects.values_list('total_nonreported_ecg', flat=True).first()

            total_reported_patients = PatientDetails.objects.filter(cardiologist__isnull=False, isDone=True).count()
            total_rejected_patients = PatientDetails.objects.filter(cardiologist__isnull=False, status=True).count()
            total_nonreported_patients = PatientDetails.objects.filter(NonReportable=True).count()
            total_unreported_and_unallocated_patients = PatientDetails.objects.filter(cardiologist=None, isDone=False).count()
            total_unreported_and_allocated_patients = PatientDetails.objects.filter(cardiologist__isnull=False, isDone=False).count()
            total_unreported_patients = total_unreported_and_unallocated_patients + total_unreported_and_allocated_patients

            total_cases = {
                'current_reported_cases': total_reported_patients,
                'total_unreported': total_unreported_patients,
                'unallocated': total_unreported_and_unallocated_patients,
                'nonreported': total_nonreported_patients 
            }

            # Get cardiologists
            cardiologist_group = Group.objects.get(name='cardiologist')
            cardiologists_objects = cardiologist_group.user_set.all()

            # Set up pagination
            paginator = Paginator(patients, 400)  # 200 patients per page
            page_number = request.GET.get('page', 1)  # Get the page number from the request
            try:
                page_obj = paginator.get_page(page_number)
            except PageNotAnInteger:
                # If page is not an integer, deliver first page
                page_obj = paginator.get_page(1)
            except EmptyPage:
                # If page is out of range, deliver last page of results
                page_obj = paginator.get_page(paginator.num_pages)

            # Get unique dates from the patients on the current page
            unique_dates = set(patient.date.date_field for patient in page_obj.object_list)
            sorted_unique_dates = sorted(unique_dates, reverse=False)
            formatted_dates = [date.strftime('%Y-%m-%d') for date in sorted_unique_dates]

            # Get unique cities and locations
            unique_cities = [f"{x.name}" for x in City.objects.all()]
            unique_locations = [f"{y.name}" for y in Location.objects.all()]

            # Everything related to the uploading of the ecg file to the database from the modal.
            form = ECGUploadForm()
            locations = Location.objects.all()
            # End of all the data of the alllocation  page.

            # Display a success message and render the same page
            messages.success(request, 'PDF uploaded and processed successfully!')
            # return render(request, 'users/allocation.html', {'form': ECGUploadForm()})

            return render(request, 'users/allocation.html', {
                'total_cases': total_cases,
                'total': total_current_uploaded,
                'count': total_uploaded_ecg,
                'total_reported': total_reported_ecg,
                'total_nonreported': total_nonreported_ecg,
                'patients': page_obj,  # Use page_obj for pagination
                'cardiologists': cardiologists_objects,
                'Date': formatted_dates,
                'Location': unique_locations,
                'Cities': unique_cities,
                'rejected': total_rejected_patients,
                'page_obj': page_obj,  # Pass page_obj for pagination controls
                'form': form,
                'location': locations,
                'success_message': success_message,
                'success_details': success_details,
                'rejected_message': rejected_message,
                'rejected_details': rejected_details,

            })

            # return render(request, 'users/allocation.html', {
            #     'success_message': success_message,
            #     'success_details': success_details,
            #     'rejected_message': rejected_message,
            #     'rejected_details': rejected_details,
            # })
            # query_params = urlencode({
            #     'success_message': success_message,
            #     'rejected_message': rejected_message,
            #     'success_details': success_details,
            #     'rejected_details': rejected_details,
            # })

            # return redirect(f"{reverse('ecgcoordinator')}?{query_params}")
        else:
            # Form is not valid, render the page with form errors
            print("Form is not valid.")
            print("Form errors:", form.errors)
            return redirect('ecgcoordinator')
    else:
        # For GET request, just render the empty form
        print("There was no post request, so redirecting to the coordinator page.")
        form = ECGUploadForm()
    
    
    return redirect('ecgcoordinator')

# End of the ECG files uploading logic on the ecg coordinator page.









